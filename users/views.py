from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import PasswordResetView
from django.utils.translation import gettext_lazy as _
from django.core.mail import send_mail
from .models import User
from .forms import UserRegisterForm
from django.db.models import Q
import logging
from departments.models import Department
from profiles.models import EmployeeProfile


logger = logging.getLogger(__name__)

def register(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get("username")
            messages.success(request, f"Account created for {username}! You can now log in.")
            return redirect("login")
    else:
        form = UserRegisterForm()
    return render(request, 'auth/register.html', {'form': form})


def employeeRegister(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get("username")
            messages.success(request, f"Account created for {username}! You can now log in.")
            return redirect("login")
    else:
        form = UserRegisterForm()
    return render(request, 'auth/employee_register.html', {'form': form})


def user_search_view(request):
    context = {}
    if request.method == "GET":
        search_query = request.GET.get("q")
        try:
            if search_query and len(search_query) > 0:
                search_results = User.objects.filter(
                    Q(username__icontains=search_query) |
                    Q(email__icontains=search_query) |
                    Q(first_name__icontains=search_query) |
                    Q(last_name__icontains=search_query)
                ).distinct()
                accounts = [(account, False) for account in search_results]
                context['accounts'] = accounts
                context['search_query'] = search_query
        except Exception as e:
            logger.error(f"Search error: {e}")
    return render(request, "users/user_search_results.html", context)





@login_required
def usersIndexView(request):
    departments = Department.objects.order_by("name")

    # Restrict access to staff or dept heads
    if request.user.is_staff or departments.filter(slug='hr').exists():
        messages.info(request, "You are seeing this page because you are a Staff/Admin or from HR Department.")
    else:
        messages.error(request, "Access Denied.")
        return redirect('home')

    # View is now based on actual assigned_department (not plantilla)
    department_users = {
        dept: EmployeeProfile.objects.filter(
            assigned_department=dept
        ).select_related('user', 'plantilla')
        for dept in departments
    }

    context_data = {
        'users': User.objects.all().order_by('last_name', 'first_name')[:50],
        'profiles': EmployeeProfile.objects.all(),
        'userCount': User.objects.count(),
        'department_users': department_users,
    }

    return render(request, 'users/users_index.html', context_data)


import csv
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from collections import defaultdict


@login_required
def export_department_users(request, department, format):
    # Look up department by slug
    dept = get_object_or_404(Department, slug=department)
    
    # Get users assigned to this department
    profiles = EmployeeProfile.objects.filter(
        assigned_department=dept
    ).select_related(
        'user', 'plantilla', 'reg_or_ct_salary', 'jo_salary'
    ).order_by('user__last_name')

    filename = f"{dept.slug}_employees.{format}"

    headers = [
        'Last Name', 'First Name', 'Ext Name', 'Username', 'Contact Number',
        'Address', 'Note', 'Employment Type', 'Date Hired (REG)', 'Date Hired (JO)',
        'Plantilla', 'Salary'
    ]

    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow(headers)
        
        for profile in profiles:
            user = profile.user
            writer.writerow([
                getattr(user, 'last_name', ""),
                getattr(user, 'first_name', ""),
                getattr(profile, 'ext_name', ""),
                getattr(user, 'username', ""),
                getattr(profile, 'contact_number', ""),
                getattr(profile, 'address', ""),
                getattr(profile, 'note', ""),
                getattr(profile, 'employment_type', ""),
                getattr(profile, 'reg_date_hired', ""),
                getattr(profile, 'jo_date_hired', ""),
                str(profile.plantilla or ""),
                profile.get_salary() or 0
            ])
        return response

    elif format == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = f"{dept.name[:30]}"
        ws.append(headers)

        for profile in profiles:
            user = profile.user
            ws.append([
                getattr(user, 'last_name', ""),
                getattr(user, 'first_name', ""),
                getattr(profile, 'ext_name', ""),
                getattr(user, 'username', ""),
                getattr(profile, 'contact_number', ""),
                getattr(profile, 'address', ""),
                getattr(profile, 'note', ""),
                getattr(profile, 'employment_type', ""),
                getattr(profile, 'reg_date_hired', ""),
                getattr(profile, 'jo_date_hired', ""),
                str(profile.plantilla or ""),
                profile.get_salary() or 0
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response

    else:
        return HttpResponse("Unsupported format", status=400)


@login_required
def export_all_employees(request, format):
    profiles = EmployeeProfile.objects.select_related(
        'user', 'assigned_department', 'plantilla',
        'reg_or_ct_salary', 'jo_salary'
    ).order_by('assigned_department__name', 'user__last_name')

    filename = f"all_employees_grouped_by_department.{format}"

    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow([
            'Department', 'Last Name', 'First Name', 'Ext Name', 'Username', 'Contact Number',
            'Address', 'Note', 'Employment Type', 'Date Hired (REG)', 'Date Hired (JO)',
            'Plantilla', 'Salary'
        ])

        for profile in profiles:
            writer.writerow([
                str(profile.assigned_department or ""),
                getattr(profile.user, 'last_name', ""),
                getattr(profile.user, 'first_name', ""),
                getattr(profile, 'ext_name', ""),
                getattr(profile.user, 'username', ""),
                getattr(profile, 'contact_number', ""),
                getattr(profile, 'address', ""),
                getattr(profile, 'note', ""),
                getattr(profile, 'employment_type', ""),
                getattr(profile, 'reg_date_hired', ""),
                getattr(profile, 'jo_date_hired', ""),
                str(profile.plantilla or ""),
                profile.get_salary() or 0
            ])
        return response

    elif format == 'excel':
        wb = Workbook()
        ws_all = wb.active
        ws_all.title = "All Employees"

        headers = [
            'Department', 'Last Name', 'First Name', 'Ext Name', 'Username', 'Contact Number',
            'Address', 'Note', 'Employment Type', 'Date Hired (REG)', 'Date Hired (JO)',
            'Plantilla', 'Salary'
        ]
        ws_all.append(headers)

        for profile in profiles:
            ws_all.append([
                str(profile.assigned_department or ""),
                getattr(profile.user, 'last_name', ""),
                getattr(profile.user, 'first_name', ""),
                getattr(profile, 'ext_name', ""),
                getattr(profile.user, 'username', ""),
                getattr(profile, 'contact_number', ""),
                getattr(profile, 'address', ""),
                getattr(profile, 'note', ""),
                getattr(profile, 'employment_type', ""),
                getattr(profile, 'reg_date_hired', ""),
                getattr(profile, 'jo_date_hired', ""),
                str(profile.plantilla or ""),
                profile.get_salary() or 0
            ])

        # Additional sheets per department
        departments = Department.objects.all().order_by('name')
        for dept in departments:
            ws = wb.create_sheet(title=dept.name[:31])
            ws.append(headers)

            dept_profiles = profiles.filter(assigned_department=dept)
            for profile in dept_profiles:
                ws.append([
                    str(profile.assigned_department or ""),
                    getattr(profile.user, 'last_name', ""),
                    getattr(profile.user, 'first_name', ""),
                    getattr(profile, 'ext_name', ""),
                    getattr(profile.user, 'username', ""),
                    getattr(profile, 'contact_number', ""),
                    getattr(profile, 'address', ""),
                    getattr(profile, 'note', ""),
                    getattr(profile, 'employment_type', ""),
                    getattr(profile, 'reg_date_hired', ""),
                    getattr(profile, 'jo_date_hired', ""),
                    str(profile.plantilla or ""),
                    profile.get_salary() or 0
                ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response

    else:
        return HttpResponse("Unsupported format", status=400)




class CustomPasswordResetView(PasswordResetView):
    email_template_name = 'registration/password_reset_email.txt'  # plain-text fallback (optional)
    html_email_template_name = 'registration/password_reset_email.html'  # ✅ your styled HTML version
    subject_template_name = 'registration/password_reset_subject.txt'  # optional
    success_url = '/password_reset/done/'  # fallback redirect after sending email

    def form_valid(self, form):
        try:
            response = super().form_valid(form)
            logger.info("Password reset email sent to: %s", form.cleaned_data['email'])
            return response
        except Exception as e:
            logger.error("Failed to send password reset email to %s: %s", form.cleaned_data['email'], str(e))
            form.add_error(None, _("There was an error sending the password reset email. Please try again later."))
            return self.form_invalid(form)
