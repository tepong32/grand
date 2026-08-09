from __future__ import annotations

from io import BytesIO
import csv

from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from collections import defaultdict

from departments.models import Department
from profiles.models import EmployeeProfile
from users.services.query_service import can_access_users_directory

from openpyxl import Workbook


PROFILE_EXPORT_HEADERS = [
    'Last Name', 'First Name', 'Ext Name', 'Username', 'Contact Number',
    'Address', 'Note', 'Employment Type', 'Date Hired (REG)', 'Date Hired (JO)',
    'Plantilla', 'Salary'
]


def _profile_to_row(profile):
    user = profile.user
    return [
        getattr(user, 'last_name', ''),
        getattr(user, 'first_name', ''),
        getattr(profile, 'ext_name', ''),
        getattr(user, 'username', ''),
        getattr(profile, 'contact_number', ''),
        getattr(profile, 'address', ''),
        getattr(profile, 'note', ''),
        getattr(profile, 'employment_type', ''),
        getattr(profile, 'reg_date_hired', ''),
        getattr(profile, 'jo_date_hired', ''),
        str(profile.plantilla or ''),
        profile.get_salary() or 0,
    ]


def _to_csv_response(filename, rows):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(PROFILE_EXPORT_HEADERS)
    for row in rows:
        writer.writerow(row)
    return response


def _to_department_excel_response(filename, sheet_name, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:30]
    ws.append(PROFILE_EXPORT_HEADERS)
    for row in rows:
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _require_export_access(actor_user):
    if not can_access_users_directory(actor_user):
        raise PermissionDenied("You do not have permission to export employee data.")


def export_department_users(department_slug, fmt, actor_user=None):
    _require_export_access(actor_user)
    dept = get_object_or_404(Department, slug=department_slug)
    profiles = EmployeeProfile.objects.filter(
        assigned_department=dept
    ).select_related(
        'user', 'plantilla', 'reg_or_ct_salary', 'jo_salary'
    ).order_by('user__last_name')

    rows = [_profile_to_row(profile) for profile in profiles]
    if fmt == 'csv':
        return _to_csv_response(f"{dept.slug}_employees.csv", rows)

    if fmt == 'excel':
        return _to_department_excel_response(f"{dept.slug}_employees.xlsx", dept.name[:30], rows)

    raise ValueError('Unsupported format')


def export_all_employees(fmt, actor_user=None):
    _require_export_access(actor_user)
    profiles = list(
        EmployeeProfile.objects.select_related(
            'user', 'assigned_department', 'plantilla',
            'reg_or_ct_salary', 'jo_salary'
        ).order_by('assigned_department__name', 'user__last_name')
    )

    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="all_employees_grouped_by_department.csv"'
        writer = csv.writer(response)
        writer.writerow(['Department'] + PROFILE_EXPORT_HEADERS)
        for profile in profiles:
            writer.writerow([str(profile.assigned_department or ''), *_profile_to_row(profile)])
        return response

    if fmt == 'excel':
        wb = Workbook()
        all_sheet = wb.active
        all_sheet.title = 'All Employees'
        all_sheet.append(['Department'] + PROFILE_EXPORT_HEADERS)

        for profile in profiles:
            all_sheet.append([str(profile.assigned_department or ''), *_profile_to_row(profile)])

        by_dept = defaultdict(list)
        for profile in profiles:
            by_dept[profile.assigned_department].append(profile)

        for dept, dept_profiles in by_dept.items():
            ws = wb.create_sheet(title=(str(dept.name) if dept else 'Unassigned')[:31])
            ws.append(PROFILE_EXPORT_HEADERS)
            for profile in dept_profiles:
                ws.append(_profile_to_row(profile))

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="all_employees_grouped_by_department.xlsx"'
        return response

    raise ValueError('Unsupported format')
