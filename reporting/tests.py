import hashlib
import io
import json
import tempfile
from datetime import timedelta
from pathlib import Path
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from pypdf import PdfReader
from reportlab.pdfgen import canvas

from assistance.models import AssistanceRequest, AssistanceType
from departments.models import Department
from social_welfare.models import ProgramActivity, SocialWelfareProgram

from .access import can_view_reporting
from .datasets import build_dataset
from .mappers import TemplateMappingError, preflight_template
from .models import ReportDefinition, ReportRun, ReportSchedule, ReportTemplateMappingField, ReportTemplateVersion
from .presets import seed_mswd_presets
from .services import create_manual_run, execute_schedule, transition_run


TEST_MEDIA_ROOT = tempfile.mkdtemp(prefix="grand-reporting-tests-")


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT)
class ReportingPlatformTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.mswd = Department.objects.create(name="Municipal Social Welfare and Development Office", slug="mswd", dashboard_template="home/authed/dashboards/mswd.html")
        cls.hr = Department.objects.create(name="Human Resources", slug="hr")
        users = get_user_model()
        cls.head = users.objects.create_user(username="report-head", email="head@example.gov", password="test-password", first_name="Mara", last_name="Santos")
        cls.operator = users.objects.create_user(username="report-operator", email="operator@example.gov", password="test-password", first_name="Leo", last_name="Cruz")
        cls.reviewer = users.objects.create_user(username="report-reviewer", email="reviewer@example.gov", password="test-password", first_name="Ana", last_name="Reyes")
        cls.limited = users.objects.create_user(username="report-limited", email="limited@example.gov", password="test-password")
        cls.outsider = users.objects.create_user(username="hr-reporter", email="hr@example.gov", password="test-password")
        for user in (cls.head, cls.operator, cls.reviewer, cls.limited):
            user.employeeprofile.assigned_department = cls.mswd
            user.employeeprofile.save(update_fields=("assigned_department",))
        cls.outsider.employeeprofile.assigned_department = cls.hr
        cls.outsider.employeeprofile.save(update_fields=("assigned_department",))
        cls.mswd.deptHead_or_oic = cls.head
        cls.mswd.save(update_fields=("deptHead_or_oic",))
        cls.operator.user_permissions.add(*Permission.objects.filter(codename__in=("view_reporting_workspace", "generate_reports", "download_reports")))
        cls.reviewer.user_permissions.add(*Permission.objects.filter(codename__in=("view_reporting_workspace", "review_reports", "approve_reports", "download_reports")))
        cls.limited.user_permissions.add(Permission.objects.get(codename="view_reporting_workspace"))
        cls.outsider.user_permissions.add(*Permission.objects.filter(codename__in=("view_reporting_workspace", "generate_reports")))

        assistance_type = AssistanceType.objects.create(name="Medical assistance", description="Support", requirements="Documents")
        cls.request = AssistanceRequest.objects.create(assistance_type=assistance_type, full_name="Synthetic Citizen", email="citizen@example.test", phone="09123456789", status="submitted")
        program = SocialWelfareProgram.objects.create(department=cls.mswd, name="Community Nutrition", code="MSWD-NUT-01", program_type="feeding", status="active", created_by=cls.head, updated_by=cls.head)
        ProgramActivity.objects.create(program=program, title="Nutrition session", activity_type="feeding", starts_at=timezone.now() - timedelta(days=2), venue="Civic Hall", status="completed", expected_attendance=60, actual_attendance=54, outcome_notes="Session completed.", created_by=cls.head, updated_by=cls.head)

        cls.definition = ReportDefinition.objects.create(department=cls.mswd, name="Assistance Volume", slug="assistance-volume", description="Volume by type and status.", dataset_key="mswd_assistance_volume", selected_fields=["assistance_type", "status", "request_count"], totals=["request_count"], default_format="pdf", created_by=cls.head, updated_by=cls.head)
        validated_at = timezone.now()
        cls.template = ReportTemplateVersion.objects.create(definition=cls.definition, version=1, title="Assistance Volume and Status", header_text="Municipal Social Welfare and Development Office", certification_text="Certified from approved operational records.", footer_text="Controlled output", document_control_prefix="MSWD-RPT", signatories=[{"role": "Prepared by", "name": "Reporting Officer"}], fidelity_status=ReportTemplateVersion.OFFICIAL, fidelity_notes="Compared with the department's current form and approved for synthetic tests.", fidelity_validated_by=cls.head, fidelity_validated_at=validated_at, created_by=cls.head, approved_by=cls.head, approved_at=validated_at)

    def _generate(self, output_format="pdf", actor=None):
        today = timezone.localdate()
        return create_manual_run(self.definition, self.template, output_format, today - timedelta(days=7), today, {}, actor or self.operator)

    def _xlsx_reference(self, rows=4):
        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Current MSWD Form"
        sheet.merge_cells("A1:C1")
        sheet["A1"] = "MUNICIPAL SOCIAL WELFARE REPORT"
        sheet["A1"].font = Font(bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor="17365D")
        sheet["A6"] = "Assistance type"
        sheet["B6"] = "Status"
        sheet["C6"] = "Requests"
        sheet["E1"] = "=SUM(C7:C10)"
        workbook.defined_names.add(DefinedName("GRAND_TITLE", attr_text="'Current MSWD Form'!$A$2"))
        workbook.defined_names.add(DefinedName("GRAND_PERIOD", attr_text="'Current MSWD Form'!$A$3"))
        workbook.defined_names.add(DefinedName("GRAND_CONTROL_ID", attr_text="'Current MSWD Form'!$A$4"))
        workbook.defined_names.add(DefinedName("GRAND_DATA_AREA", attr_text=f"'Current MSWD Form'!$A$7:$C${6 + rows}"))
        workbook.defined_names.add(DefinedName("GRAND_TOTALS_AREA", attr_text=f"'Current MSWD Form'!$A${7 + rows}:$C${7 + rows}"))
        workbook.save(output)
        return output.getvalue()

    def _pdf_reference(self):
        output = io.BytesIO()
        document = canvas.Canvas(output)
        document.setFont("Helvetica-Bold", 12)
        document.drawString(40, 800, "Existing MSWD Accomplishment Form")
        document.rect(35, 650, 520, 120)
        document.save()
        return output.getvalue()

    def test_department_head_and_explicit_permission_can_access_workspace(self):
        self.assertTrue(can_view_reporting(self.head))
        self.assertTrue(can_view_reporting(self.operator))
        self.client.force_login(self.operator)
        response = self.client.get(reverse("reporting:workspace"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.definition.name)
        self.assertContains(response, "Uploaded forms are references only")

    def test_department_boundary_prevents_cross_department_report_access(self):
        self.client.force_login(self.outsider)
        response = self.client.get(self.definition.get_absolute_url())
        self.assertEqual(response.status_code, 404)
        workspace = self.client.get(reverse("reporting:workspace"))
        self.assertEqual(workspace.status_code, 200)
        self.assertNotContains(workspace, self.definition.name)

    def test_mswd_dataset_cannot_be_configured_for_another_department(self):
        definition = ReportDefinition(department=self.hr, name="Cross-boundary", slug="cross-boundary", dataset_key="mswd_assistance_volume", selected_fields=["assistance_type", "request_count"], created_by=self.outsider, updated_by=self.outsider)
        with self.assertRaises(ValidationError):
            definition.full_clean()

    def test_definition_rejects_unknown_dataset_fields_and_executable_configuration(self):
        definition = ReportDefinition(department=self.mswd, name="Unsafe", slug="unsafe", dataset_key="mswd_assistance_volume", selected_fields=["raw_sql"], filters="SELECT *", created_by=self.head, updated_by=self.head)
        with self.assertRaises(ValidationError):
            definition.full_clean()

    def test_controlled_filter_group_total_and_sort_configuration_is_applied(self):
        grouped = ReportDefinition(
            department=self.mswd, name="Submitted assistance by type", slug="submitted-by-type",
            dataset_key="mswd_assistance_volume", selected_fields=["assistance_type", "request_count"],
            filters={"status__exact": "Submitted"}, group_by=["assistance_type"], totals=["request_count"],
            sort_by=["-request_count"], created_by=self.head, updated_by=self.head,
        )
        grouped.full_clean()
        adapter, rows, totals = build_dataset(grouped, timezone.localdate() - timedelta(days=7), timezone.localdate(), {})
        self.assertEqual(adapter.key, "mswd_assistance_volume")
        self.assertEqual(rows, [{"assistance_type": "Medical assistance", "request_count": 1}])
        self.assertEqual(totals, {"request_count": 1})

    def test_pdf_generation_archives_checksum_and_audit_event(self):
        run = self._generate("pdf")
        payload = Path(run.output_file.path).read_bytes()
        self.assertTrue(payload.startswith(b"%PDF"))
        self.assertEqual(run.status, ReportRun.GENERATED)
        self.assertEqual(run.checksum, hashlib.sha256(payload).hexdigest())
        self.assertEqual(run.row_count, 1)
        self.assertIn(f"_{run.period_end:%Y%m%d}_", run.output_file.name)
        self.assertEqual(run.parameters["_definition_snapshot"]["selected_fields"], self.definition.selected_fields)
        self.assertEqual(run.events.get().action, "generated")

    def test_workspace_permission_alone_does_not_expose_colleagues_report_runs(self):
        run = self._generate("csv", self.operator)
        self.client.force_login(self.limited)
        workspace = self.client.get(reverse("reporting:workspace"))
        self.assertNotContains(workspace, run.get_absolute_url())
        self.assertEqual(self.client.get(run.get_absolute_url()).status_code, 404)

    def test_xlsx_and_csv_outputs_are_structured_and_readable(self):
        xlsx_run = self._generate("xlsx")
        workbook = load_workbook(xlsx_run.output_file.path, data_only=False)
        sheet = workbook["Official Report"]
        self.assertEqual(sheet["A2"].value, self.template.title)
        self.assertEqual(sheet["A6"].value, "Assistance type")
        self.assertEqual(sheet["C7"].value, 1)
        self.assertEqual(sheet.page_setup.orientation, "landscape")
        self.assertEqual(sheet.print_title_rows, "$6:$6")
        self.assertIn("Page &P of &N", sheet.oddFooter.center.text)
        csv_run = self._generate("csv")
        text = Path(csv_run.output_file.path).read_text(encoding="utf-8-sig")
        self.assertIn("Document control", text)
        self.assertIn("Medical assistance,Submitted,1", text)

    def test_generated_report_requires_review_before_approval(self):
        run = self._generate()
        with self.assertRaises(ValueError):
            transition_run(run, "approve", self.reviewer)
        transition_run(run, "review", self.reviewer, "Figures checked.")
        transition_run(run, "approve", self.reviewer, "Approved for filing.")
        run.refresh_from_db()
        self.assertEqual(run.status, ReportRun.APPROVED)
        self.assertEqual(run.events.count(), 3)

    def test_pilot_template_blocks_official_approval_until_fidelity_validation(self):
        pilot = ReportTemplateVersion.objects.create(
            definition=self.definition, version=2, title="Current office form pilot",
            created_by=self.head, approved_by=self.head, approved_at=timezone.now(),
        )
        today = timezone.localdate()
        run = create_manual_run(self.definition, pilot, "pdf", today, today, {}, self.operator)
        transition_run(run, "review", self.reviewer)
        with self.assertRaisesMessage(ValueError, "pilot layout"):
            transition_run(run, "approve", self.reviewer)
        run.refresh_from_db()
        self.assertEqual(run.status, ReportRun.REVIEWED)

    def test_direct_fidelity_shortcut_is_retired_in_favor_of_promotion_record(self):
        pilot = ReportTemplateVersion.objects.create(
            definition=self.definition, version=2, title="Validated office form",
            created_by=self.head, approved_by=self.head, approved_at=timezone.now(),
        )
        self.client.force_login(self.reviewer)
        url = reverse("reporting:template_validate_fidelity", args=(pilot.pk,))
        response = self.client.post(url, {"fidelity_notes": "A free-text note is no longer enough."})
        self.assertRedirects(
            response, reverse("reporting:template_promotion_create", args=(pilot.pk,)),
            fetch_redirect_response=False,
        )
        pilot.refresh_from_db()
        self.assertFalse(pilot.is_official_ready)

    def test_legacy_approval_without_fidelity_is_not_presented_as_current_official_output(self):
        pilot = ReportTemplateVersion.objects.create(
            definition=self.definition, version=2, title="Earlier approved form",
            created_by=self.head, approved_by=self.head, approved_at=timezone.now(),
        )
        today = timezone.localdate()
        run = ReportRun.objects.create(
            definition=self.definition, template_version=pilot, output_format="pdf", period_start=today,
            period_end=today, parameters={}, idempotency_key="legacy:approved-pilot", status=ReportRun.APPROVED,
            created_by=self.operator, approved_by=self.reviewer, approved_at=timezone.now(),
        )
        self.assertFalse(run.is_official_output)
        self.client.force_login(self.operator)
        detail = self.client.get(run.get_absolute_url())
        self.assertContains(detail, "Earlier approval retained")
        self.assertNotContains(detail, "Approved for official use")
        workspace = self.client.get(reverse("reporting:workspace"))
        self.assertContains(workspace, "No official outputs approved yet")

    def test_new_approval_supersedes_same_period_without_deleting_prior_output(self):
        first = self._generate()
        transition_run(first, "review", self.reviewer)
        transition_run(first, "approve", self.reviewer)
        second = self._generate()
        second.period_start, second.period_end = first.period_start, first.period_end
        second.save(update_fields=("period_start", "period_end"))
        transition_run(second, "review", self.reviewer)
        transition_run(second, "approve", self.reviewer)
        first.refresh_from_db()
        self.assertEqual(first.status, ReportRun.SUPERSEDED)
        self.assertTrue(first.events.filter(action="superseded_by_new_approval").exists())
        self.assertTrue(first.output_file.storage.exists(first.output_file.name))

    def test_schedule_ledger_is_idempotent_and_advances_once(self):
        due = timezone.now() - timedelta(minutes=5)
        schedule = ReportSchedule.objects.create(definition=self.definition, template_version=self.template, name="Daily assistance report", frequency="daily", output_format="csv", next_run_at=due, created_by=self.head)
        first, first_created = execute_schedule(schedule, due)
        schedule.refresh_from_db()
        advanced_to = schedule.next_run_at
        second, second_created = execute_schedule(schedule, due)
        schedule.refresh_from_db()
        self.assertTrue(first_created)
        self.assertFalse(second_created)
        self.assertEqual(first.pk, second.pk)
        self.assertEqual(schedule.next_run_at, advanced_to)

    def test_failed_run_records_error_without_losing_prior_success(self):
        successful = self._generate()
        today = timezone.localdate()
        run = ReportRun.objects.create(definition=self.definition, template_version=self.template, output_format="pdf", period_start=today, period_end=today, parameters={}, idempotency_key="manual:forced-failure", created_by=self.operator)
        with patch("reporting.services.build_dataset", side_effect=RuntimeError("Synthetic adapter failure")):
            from .services import generate_report
            with self.assertRaises(RuntimeError):
                generate_report(run)
        run.refresh_from_db()
        self.assertEqual(run.status, ReportRun.FAILED)
        self.assertTrue(successful.output_file.storage.exists(successful.output_file.name))

    def test_scheduled_failure_persists_and_safe_rerun_reuses_ledger_entry(self):
        due = timezone.now() - timedelta(minutes=5)
        schedule = ReportSchedule.objects.create(definition=self.definition, template_version=self.template, name="Retryable schedule", frequency="daily", output_format="csv", next_run_at=due, created_by=self.head)
        with patch("reporting.services.build_dataset", side_effect=RuntimeError("Temporary source failure")):
            with self.assertRaises(RuntimeError):
                execute_schedule(schedule, due)
        failed = ReportRun.objects.get(schedule=schedule)
        self.assertEqual(failed.status, ReportRun.FAILED)
        schedule.refresh_from_db()
        self.assertEqual(schedule.next_run_at, due)
        recovered, created = execute_schedule(schedule, due)
        self.assertFalse(created)
        self.assertEqual(recovered.pk, failed.pk)
        self.assertEqual(recovered.status, ReportRun.GENERATED)

    def test_uploaded_reference_rejects_executable_file_extension(self):
        reference = ReportTemplateVersion(definition=self.definition, version=2, title="Unsafe reference", reference_kind="docx", reference_file=SimpleUploadedFile("macro.py", b"print('no')"), created_by=self.head)
        with self.assertRaises(ValidationError):
            reference.full_clean()

    def test_approved_template_version_cannot_be_edited_in_place(self):
        template = ReportTemplateVersion.objects.get(pk=self.template.pk)
        template.title = "Changed after approval"
        with self.assertRaises(ValidationError):
            template.full_clean()

    def test_approved_template_print_geometry_cannot_be_edited_in_place(self):
        template = ReportTemplateVersion.objects.get(pk=self.template.pk)
        template.margin_mm = 22
        with self.assertRaises(ValidationError):
            template.full_clean()

    def test_preset_seeding_is_repeatable_and_creates_five_approved_reports(self):
        self.definition.delete()
        first = seed_mswd_presets(self.head)
        second = seed_mswd_presets(self.head)
        self.assertEqual(len(first), 5)
        self.assertEqual(len(second), 5)
        self.assertEqual(ReportDefinition.objects.filter(department=self.mswd).count(), 5)
        self.assertEqual(ReportTemplateVersion.objects.filter(definition__department=self.mswd, approved_at__isnull=False).count(), 5)

    def test_mswd_dashboard_activates_reporting_only_for_authorized_employee(self):
        self.client.force_login(self.operator)
        response = self.client.get(reverse("department_dashboard"))
        section = response.context["dashboard_sections"][5]
        self.assertEqual(section["status"], "Available")
        self.assertContains(response, "Open Reporting Workspace")
        self.assertContains(response, reverse("reporting:workspace"))

    def test_download_requires_separate_permission(self):
        run = self._generate("csv")
        self.client.force_login(self.operator)
        with tempfile.TemporaryDirectory() as export_root, self.settings(GRAND_EXPORT_ROOT=export_root):
            response = self.client.get(reverse("reporting:run_download", args=(run.public_id,)))
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response["X-GRAND-Export-Archived"], "true")
            artifacts = list(Path(export_root).rglob("*.csv"))
            self.assertEqual(len(artifacts), 1)
            self.assertIn(self.mswd.slug, artifacts[0].parts)
            self.assertIn(self.operator.username, artifacts[0].parts)
            manifest = json.loads(Path(str(artifacts[0]) + ".manifest.json").read_text(encoding="utf-8"))
            self.assertEqual(manifest["metadata"]["run_public_id"], str(run.public_id))
            self.assertEqual(manifest["sha256"], response["X-GRAND-Export-SHA256"])
            self.assertTrue(run.events.filter(action="exported", actor=self.operator).exists())
        self.operator.user_permissions.remove(Permission.objects.get(codename="download_reports"))
        self.operator = get_user_model().objects.get(pk=self.operator.pk)
        self.client.force_login(self.operator)
        self.assertEqual(self.client.get(reverse("reporting:run_download", args=(run.public_id,))).status_code, 403)

    def test_print_action_is_visible_only_for_authorized_pdf_outputs(self):
        pdf_run = self._generate("pdf")
        xlsx_run = self._generate("xlsx")
        self.client.force_login(self.operator)
        pdf_detail = self.client.get(pdf_run.get_absolute_url())
        self.assertContains(pdf_detail, "Print PDF")
        preview = self.client.get(reverse("reporting:run_print_preview", args=(pdf_run.public_id,)))
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview["Content-Type"], "application/pdf")
        self.assertIn("inline", preview["Content-Disposition"])
        xlsx_detail = self.client.get(xlsx_run.get_absolute_url())
        self.assertNotContains(xlsx_detail, "Print PDF")
        self.assertEqual(self.client.get(reverse("reporting:run_print_preview", args=(xlsx_run.public_id,))).status_code, 404)

    def test_print_and_download_actions_are_hidden_without_download_permission(self):
        run = self._generate("pdf")
        self.client.force_login(self.limited)
        self.assertEqual(self.client.get(run.get_absolute_url()).status_code, 404)
        self.client.force_login(self.operator)
        self.operator.user_permissions.remove(Permission.objects.get(codename="download_reports"))
        self.operator = get_user_model().objects.get(pk=self.operator.pk)
        self.client.force_login(self.operator)
        detail = self.client.get(run.get_absolute_url())
        self.assertNotContains(detail, "Print PDF")
        self.assertNotContains(detail, "Download PDF")
        self.assertContains(detail, "Printing and downloading require report download permission")
        self.assertEqual(self.client.get(reverse("reporting:run_print_preview", args=(run.public_id,))).status_code, 403)

    def test_mapped_xlsx_preserves_department_layout_and_populates_reserved_ranges(self):
        template = ReportTemplateVersion.objects.create(
            definition=self.definition, version=2, title="Current MSWD spreadsheet",
            render_mode=ReportTemplateVersion.RENDER_XLSX_TEMPLATE,
            reference_kind=ReportTemplateVersion.REFERENCE_XLSX,
            reference_file=SimpleUploadedFile("current-mswd.xlsx", self._xlsx_reference()), created_by=self.head,
        )
        summary = preflight_template(template, self.head)
        template.refresh_from_db()
        self.assertEqual(summary["row_capacity"], 4)
        template.approved_by = self.head
        template.approved_at = timezone.now()
        template.full_clean()
        template.save(update_fields=("approved_by", "approved_at"))
        today = timezone.localdate()
        run = create_manual_run(self.definition, template, "xlsx", today - timedelta(days=7), today, {}, self.operator)
        workbook = load_workbook(run.output_file.path, data_only=False)
        sheet = workbook["Current MSWD Form"]
        self.assertEqual(sheet["A1"].value, "MUNICIPAL SOCIAL WELFARE REPORT")
        self.assertEqual(sheet["A1"].fill.fgColor.rgb, "0017365D")
        self.assertEqual(sheet["A2"].value, template.title)
        self.assertEqual(sheet["A7"].value, "Medical assistance")
        self.assertEqual(sheet["C7"].value, 1)
        self.assertEqual(sheet["E1"].value, "=SUM(C7:C10)")
        self.assertEqual(run.parameters["_template_snapshot"]["mapping_checksum"], template.mapping_checksum)

    def test_mapped_xlsx_rejects_wrong_output_and_capacity_overflow(self):
        template = ReportTemplateVersion.objects.create(
            definition=self.definition, version=2, title="One-row form",
            render_mode=ReportTemplateVersion.RENDER_XLSX_TEMPLATE,
            reference_kind=ReportTemplateVersion.REFERENCE_XLSX,
            reference_file=SimpleUploadedFile("one-row.xlsx", self._xlsx_reference(rows=1)), created_by=self.head,
        )
        preflight_template(template, self.head)
        template.refresh_from_db()
        template.approved_by, template.approved_at = self.head, timezone.now()
        template.full_clean()
        template.save(update_fields=("approved_by", "approved_at"))
        today = timezone.localdate()
        with self.assertRaisesMessage(ValueError, "does not support"):
            create_manual_run(self.definition, template, "pdf", today, today, {}, self.operator)
        from .datasets import dataset_registry
        adapter = dataset_registry[self.definition.dataset_key]
        with patch("reporting.services.build_dataset", return_value=(adapter, [{"assistance_type": "A", "status": "Submitted", "request_count": 1}, {"assistance_type": "B", "status": "Pending", "request_count": 1}], {"request_count": 2})):
            with self.assertRaisesMessage(TemplateMappingError, "reserves 1 row"):
                create_manual_run(self.definition, template, "xlsx", today, today, {}, self.operator)
        self.assertEqual(ReportRun.objects.latest("created_at").status, ReportRun.FAILED)

    def test_pdf_overlay_keeps_original_form_and_adds_controlled_values(self):
        template = ReportTemplateVersion.objects.create(
            definition=self.definition, version=2, title="Existing PDF form",
            render_mode=ReportTemplateVersion.RENDER_PDF_OVERLAY,
            reference_kind=ReportTemplateVersion.REFERENCE_PDF,
            reference_file=SimpleUploadedFile("existing-form.pdf", self._pdf_reference()), created_by=self.head,
        )
        ReportTemplateMappingField.objects.create(template_version=template, source_key="title", page_number=1, x_mm=15, y_mm=25, width_mm=120)
        ReportTemplateMappingField.objects.create(template_version=template, source_key="assistance_type", page_number=1, x_mm=15, y_mm=45, width_mm=70, repeat_for_rows=True, max_rows=10)
        ReportTemplateMappingField.objects.create(template_version=template, source_key="request_count", page_number=1, x_mm=150, y_mm=45, width_mm=25, alignment="right", repeat_for_rows=True, max_rows=10)
        preflight_template(template, self.head)
        template.refresh_from_db()
        template.approved_by, template.approved_at = self.head, timezone.now()
        template.full_clean()
        template.save(update_fields=("approved_by", "approved_at"))
        today = timezone.localdate()
        run = create_manual_run(self.definition, template, "pdf", today - timedelta(days=7), today, {}, self.operator)
        text = "\n".join(page.extract_text() or "" for page in PdfReader(run.output_file.path).pages)
        self.assertIn("Existing MSWD Accomplishment Form", text)
        self.assertIn("Existing PDF form", text)
        self.assertIn("Medical assistance", text)

    def test_mapping_changes_invalidate_preflight_and_approved_mapping_is_locked(self):
        template = ReportTemplateVersion.objects.create(
            definition=self.definition, version=2, title="Governed PDF",
            render_mode=ReportTemplateVersion.RENDER_PDF_OVERLAY,
            reference_kind=ReportTemplateVersion.REFERENCE_PDF,
            reference_file=SimpleUploadedFile("governed.pdf", self._pdf_reference()), created_by=self.head,
        )
        mapping = ReportTemplateMappingField.objects.create(template_version=template, source_key="title", page_number=1, x_mm=15, y_mm=25, width_mm=120)
        preflight_template(template, self.head)
        mapping.x_mm = 20
        mapping.save()
        template.refresh_from_db()
        self.assertFalse(template.is_mapping_ready)
        preflight_template(template, self.head)
        template.refresh_from_db()
        template.approved_by, template.approved_at = self.head, timezone.now()
        template.full_clean()
        template.save(update_fields=("approved_by", "approved_at"))
        mapping.refresh_from_db()
        mapping.x_mm = 25
        with self.assertRaisesMessage(ValidationError, "immutable"):
            mapping.save()

    def test_mapper_workspace_obeys_department_and_management_permissions(self):
        template = ReportTemplateVersion.objects.create(
            definition=self.definition, version=2, title="Managed workbook",
            render_mode=ReportTemplateVersion.RENDER_XLSX_TEMPLATE,
            reference_kind=ReportTemplateVersion.REFERENCE_XLSX,
            reference_file=SimpleUploadedFile("managed.xlsx", self._xlsx_reference()), created_by=self.head,
        )
        self.client.force_login(self.head)
        response = self.client.get(reverse("reporting:template_mapping", args=(template.pk,)))
        self.assertContains(response, "GRAND_DATA_AREA")
        self.outsider.user_permissions.add(Permission.objects.get(codename="manage_report_templates"))
        self.outsider = get_user_model().objects.get(pk=self.outsider.pk)
        self.client.force_login(self.outsider)
        self.assertEqual(self.client.get(reverse("reporting:template_mapping", args=(template.pk,))).status_code, 404)
        self.client.force_login(self.operator)
        self.assertEqual(self.client.get(reverse("reporting:template_mapping", args=(template.pk,))).status_code, 403)
