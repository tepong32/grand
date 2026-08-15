from __future__ import annotations

import csv
import hashlib
import io
import uuid
from datetime import date, timedelta

from dateutil.relativedelta import relativedelta
from django.core.files.base import ContentFile
from django.db import transaction
from django.utils import timezone
from django.utils.formats import date_format
from openpyxl import Workbook
from openpyxl.drawing.image import Image as SpreadsheetImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, LEGAL, LETTER, landscape, portrait
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.utils import ImageReader

from .datasets import build_dataset
from .mappers import generate_mapped_xlsx, generate_pdf_overlay
from .models import ReportDefinition, ReportRun, ReportRunEvent, ReportSchedule, ReportTemplateVersion


def display_value(value):
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        if hasattr(value, "hour"):
            return date_format(timezone.localtime(value), "M j, Y g:i A")
        return value.isoformat()
    return str(value)


def definition_snapshot(definition):
    return {
        "dataset_key": definition.dataset_key,
        "selected_fields": list(definition.selected_fields),
        "filters": dict(definition.filters or {}),
        "group_by": list(definition.group_by or []),
        "totals": list(definition.totals or []),
        "sort_by": list(definition.sort_by or []),
    }


def run_parameters(definition, runtime_parameters=None, template_version=None):
    parameters = dict(runtime_parameters or {})
    parameters["_definition_snapshot"] = definition_snapshot(definition)
    if template_version:
        parameters["_template_snapshot"] = {
            "version": template_version.version,
            "render_mode": template_version.render_mode,
            "mapping_checksum": template_version.mapping_checksum,
            "mapping_summary": dict(template_version.mapping_summary or {}),
        }
    return parameters


def _selected_fields(run):
    return run.parameters.get("_definition_snapshot", {}).get("selected_fields", run.definition.selected_fields)


def _document_metadata(run):
    template = run.template_version
    prefix = template.document_control_prefix or run.definition.department.slug.upper()
    return {
        "title": template.title or run.definition.name,
        "header": template.header_text or run.definition.department.name,
        "period": f"{run.period_start:%B %d, %Y} to {run.period_end:%B %d, %Y}",
        "control_id": f"{prefix}-{str(run.public_id)[:8].upper()}",
    }


def _page_layout(template):
    sizes = {
        template.PAGE_A4: A4,
        template.PAGE_LETTER: LETTER,
        template.PAGE_LEGAL: LEGAL,
    }
    base_size = sizes.get(template.page_size, A4)
    page_size = portrait(base_size) if template.orientation == template.PORTRAIT else landscape(base_size)
    return page_size, template.margin_mm * mm


def _stored_image_path(field):
    if not field:
        return None
    try:
        return field.path
    except (NotImplementedError, ValueError):
        return None


def _generate_csv(run, labels, rows):
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    metadata = _document_metadata(run)
    writer.writerow([metadata["title"]])
    writer.writerow([metadata["header"]])
    writer.writerow(["Covered period", metadata["period"]])
    writer.writerow(["Document control", metadata["control_id"]])
    writer.writerow([])
    writer.writerow(labels)
    for row in rows:
        writer.writerow([display_value(row[key]) for key in _selected_fields(run)])
    return output.getvalue().encode("utf-8-sig")


def _generate_xlsx(run, labels, rows, totals):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Official Report"
    metadata = _document_metadata(run)
    template = run.template_version
    column_count = max(len(labels), 1)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    sheet["A1"] = metadata["header"]
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=column_count)
    sheet["A2"] = metadata["title"]
    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=column_count)
    sheet["A3"] = f"Covered period: {metadata['period']}"
    sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=column_count)
    sheet["A4"] = f"Document control: {metadata['control_id']}"
    for cell in (sheet["A1"], sheet["A2"], sheet["A3"], sheet["A4"]):
        cell.alignment = Alignment(horizontal="center")
    sheet["A1"].font = Font(bold=True, size=13)
    sheet["A2"].font = Font(bold=True, size=16, color="17365D")
    paper_sizes = {template.PAGE_A4: sheet.PAPERSIZE_A4, template.PAGE_LETTER: sheet.PAPERSIZE_LETTER, template.PAGE_LEGAL: sheet.PAPERSIZE_LEGAL}
    sheet.page_setup.paperSize = paper_sizes.get(template.page_size, sheet.PAPERSIZE_A4)
    sheet.page_setup.orientation = template.orientation
    margin_inches = template.margin_mm / 25.4
    sheet.page_margins.left = sheet.page_margins.right = margin_inches
    sheet.page_margins.top = sheet.page_margins.bottom = margin_inches
    header_row = 6
    navy = "17365D"
    for index, label in enumerate(labels, 1):
        cell = sheet.cell(header_row, index, label)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="C8D1DB")
    for row_index, row in enumerate(rows, header_row + 1):
        for column_index, key in enumerate(_selected_fields(run), 1):
            value = row[key]
            cell = sheet.cell(row_index, column_index, value)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if hasattr(value, "tzinfo"):
                cell.value = timezone.localtime(value).replace(tzinfo=None)
                cell.number_format = "mmm d, yyyy h:mm AM/PM"
            elif isinstance(value, date):
                cell.number_format = "mmm d, yyyy"
    if totals:
        totals_row = header_row + len(rows) + 1
        sheet.cell(totals_row, 1, "Totals").font = Font(bold=True)
        for key, value in totals.items():
            if key in _selected_fields(run):
                cell = sheet.cell(totals_row, _selected_fields(run).index(key) + 1, value)
                cell.font = Font(bold=True)
    sheet.freeze_panes = f"A{header_row + 1}"
    if template.repeat_header:
        sheet.print_title_rows = f"{header_row}:{header_row}"
    sheet.auto_filter.ref = f"A{header_row}:{sheet.cell(header_row + len(rows), column_count).coordinate}"
    for column_index, column in enumerate(sheet.columns, 1):
        letter = get_column_letter(column_index)
        content_width = max((len(str(cell.value or "")) for cell in column), default=10)
        sheet.column_dimensions[letter].width = min(max(content_width + 2, 12), 42)
    signatory_row = header_row + len(rows) + (3 if totals else 2)
    for signatory in run.template_version.signatories or []:
        sheet.cell(signatory_row, 1, signatory.get("role", "Prepared by"))
        sheet.cell(signatory_row + 2, 1, signatory.get("name", "____________________________"))
        signatory_row += 4
    for image_field, anchor in ((template.primary_logo, "A1"), (template.secondary_logo, get_column_letter(column_count) + "1")):
        image_path = _stored_image_path(image_field)
        if image_path:
            image = SpreadsheetImage(image_path)
            image.height = 42
            image.width = 42
            sheet.add_image(image, anchor)
    footer_parts = []
    if template.show_footer and template.footer_text:
        footer_parts.append(template.footer_text)
    if template.show_page_numbers:
        footer_parts.append("Page &P of &N")
    if template.show_document_control:
        footer_parts.append(metadata["control_id"])
    sheet.oddFooter.center.text = " | ".join(footer_parts)
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _generate_pdf(run, labels, rows, totals):
    output = io.BytesIO()
    template = run.template_version
    page_size, margin = _page_layout(template)
    metadata = _document_metadata(run)
    has_logo = bool(template.primary_logo or template.secondary_logo)
    header_band = 20 * mm if template.repeat_header or has_logo else 0
    footer_band = 8 * mm if template.show_footer or template.show_page_numbers or template.show_document_control else 0
    document = SimpleDocTemplate(output, pagesize=page_size, leftMargin=margin, rightMargin=margin, topMargin=margin + header_band, bottomMargin=margin + footer_band, title=metadata["title"], author=run.definition.department.name)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("OfficialTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=15, leading=18, alignment=TA_CENTER, textColor=colors.HexColor("#17365D"), spaceAfter=5)
    center_style = ParagraphStyle("Center", parent=styles["Normal"], alignment=TA_CENTER, fontSize=9, leading=12)
    cell_style = ParagraphStyle("Cell", parent=styles["BodyText"], fontSize=7.5, leading=9)
    header_cell_style = ParagraphStyle("HeaderCell", parent=cell_style, fontName="Helvetica-Bold", textColor=colors.white)
    story = []
    if not template.repeat_header:
        story.append(Paragraph(metadata["header"], center_style))
    story.extend([Paragraph(metadata["title"], title_style), Paragraph(f"Covered period: {metadata['period']}" + (f" &nbsp;&nbsp; | &nbsp;&nbsp; Document control: {metadata['control_id']}" if template.show_document_control else ""), center_style), Spacer(1, 7 * mm)])
    data = [[Paragraph(str(label), header_cell_style) for label in labels]]
    for row in rows:
        data.append([Paragraph(display_value(row[key]).replace("&", "&amp;").replace("<", "&lt;"), cell_style) for key in _selected_fields(run)])
    if totals:
        total_row = []
        for index, key in enumerate(_selected_fields(run)):
            total_row.append(Paragraph("Totals" if index == 0 else display_value(totals.get(key, "")), cell_style))
        data.append(total_row)
    usable_width = page_size[0] - (2 * margin)
    widths = [usable_width / max(len(labels), 1)] * max(len(labels), 1)
    table = Table(data, colWidths=widths, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365D")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B8C3CE")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F6F8")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4), ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.extend([table, Spacer(1, 9 * mm)])
    if run.template_version.certification_text:
        story.extend([Paragraph(run.template_version.certification_text, styles["BodyText"]), Spacer(1, 8 * mm)])
    for signatory in run.template_version.signatories or []:
        story.append(Paragraph(f"{signatory.get('role', 'Prepared by')}: ____________________________ &nbsp;&nbsp; {signatory.get('name', '')}", styles["BodyText"]))
        story.append(Spacer(1, 4 * mm))

    def decorate_page(canvas, doc):
        canvas.saveState()
        width, height = page_size
        if template.page_border == template.BORDER_SINGLE:
            canvas.setStrokeColor(colors.HexColor("#65717C"))
            canvas.setLineWidth(0.6)
            canvas.rect(margin / 2, margin / 2, width - margin, height - margin)
        if template.repeat_header:
            canvas.setFont("Helvetica-Bold", 9)
            canvas.setFillColor(colors.HexColor("#23384D"))
            canvas.drawCentredString(width / 2, height - margin - 4 * mm, metadata["header"])
        logo_y = height - margin - 13 * mm
        for image_field, x in ((template.primary_logo, margin), (template.secondary_logo, width - margin - 12 * mm)):
            image_path = _stored_image_path(image_field)
            if image_path:
                canvas.drawImage(ImageReader(image_path), x, logo_y, width=12 * mm, height=12 * mm, preserveAspectRatio=True, mask="auto")
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#4F5B66"))
        if template.show_footer and template.footer_text:
            canvas.drawString(margin, margin, template.footer_text)
        right_parts = []
        if template.show_page_numbers:
            right_parts.append(f"Page {doc.page}")
        if template.show_document_control:
            right_parts.append(metadata["control_id"])
        if right_parts:
            canvas.drawRightString(width - margin, margin, " | ".join(right_parts))
        canvas.restoreState()

    document.build(story, onFirstPage=decorate_page, onLaterPages=decorate_page)
    return output.getvalue()


GENERATORS = {ReportDefinition.FORMAT_CSV: _generate_csv, ReportDefinition.FORMAT_XLSX: _generate_xlsx, ReportDefinition.FORMAT_PDF: _generate_pdf}


def generate_report(run):
    if run.status not in (ReportRun.DRAFT, ReportRun.FAILED):
        return run
    try:
        adapter, rows, totals = build_dataset(run.definition, run.period_start, run.period_end, run.parameters)
        labels = adapter.labels_for(_selected_fields(run))
        template = run.template_version
        if not template.supports_format(run.output_format):
            raise ValueError("The selected template does not support this output format.")
        if not template.is_mapping_ready:
            raise ValueError("The selected mapped template has not passed preflight.")
        if template.render_mode == ReportTemplateVersion.RENDER_XLSX_TEMPLATE:
            payload = generate_mapped_xlsx(run, _document_metadata(run), rows, totals, _selected_fields(run), display_value)
        elif template.render_mode == ReportTemplateVersion.RENDER_PDF_OVERLAY:
            payload = generate_pdf_overlay(run, _document_metadata(run), rows, totals, display_value)
        elif run.output_format == ReportDefinition.FORMAT_CSV:
            payload = _generate_csv(run, labels, rows)
        else:
            payload = GENERATORS[run.output_format](run, labels, rows, totals)
        checksum = hashlib.sha256(payload).hexdigest()
        filename = f"{run.definition.slug}_{run.period_start:%Y%m%d}_{run.period_end:%Y%m%d}_{str(run.public_id)[:8]}.{run.output_format}"
        previous = run.status
        run.output_file.save(filename, ContentFile(payload), save=False)
        run.status = ReportRun.GENERATED
        run.checksum = checksum
        run.row_count = len(rows)
        run.error_message = ""
        run.generated_at = timezone.now()
        run.full_clean()
        run.save()
        ReportRunEvent.objects.create(run=run, actor=run.created_by, action="generated", from_status=previous, to_status=run.status, note=f"Generated {len(rows)} data rows with SHA-256 {checksum}.")
    except Exception as exc:
        previous = run.status
        run.status = ReportRun.FAILED
        run.error_message = str(exc)[:4000]
        run.save(update_fields=("status", "error_message", "updated_at"))
        ReportRunEvent.objects.create(run=run, actor=run.created_by, action="generation_failed", from_status=previous, to_status=run.status, note=run.error_message)
        raise
    return run


def create_manual_run(definition, template_version, output_format, period_start, period_end, parameters, actor):
    if not template_version.supports_format(output_format):
        raise ValueError("The selected template does not support this output format.")
    if not template_version.is_mapping_ready:
        raise ValueError("The selected mapped template has not passed preflight.")
    run = ReportRun(
        definition=definition, template_version=template_version, output_format=output_format,
        period_start=period_start, period_end=period_end, parameters=run_parameters(definition, parameters, template_version),
        idempotency_key=f"manual:{uuid.uuid4()}", created_by=actor,
    )
    run.full_clean()
    run.save()
    return generate_report(run)


@transaction.atomic
def transition_run(run, action, actor, note=""):
    previous = run.status
    now = timezone.now()
    if action == "review" and run.status == ReportRun.GENERATED:
        run.status, run.reviewed_by, run.reviewed_at = ReportRun.REVIEWED, actor, now
    elif action == "approve" and run.status == ReportRun.REVIEWED:
        if not run.template_version.is_official_ready:
            raise ValueError("This output uses a pilot layout. Validate the template against the department's current form before approving it as official.")
        previous_approved = list(ReportRun.objects.filter(definition=run.definition, period_start=run.period_start, period_end=run.period_end, status=ReportRun.APPROVED).exclude(pk=run.pk))
        for prior in previous_approved:
            prior.status = ReportRun.SUPERSEDED
            prior.save(update_fields=("status", "updated_at"))
            ReportRunEvent.objects.create(run=prior, actor=actor, action="superseded_by_new_approval", from_status=ReportRun.APPROVED, to_status=ReportRun.SUPERSEDED, note=f"Superseded by report run {run.public_id}.")
        run.status, run.approved_by, run.approved_at = ReportRun.APPROVED, actor, now
    elif action == "supersede" and run.status == ReportRun.APPROVED:
        run.status = ReportRun.SUPERSEDED
    else:
        raise ValueError("That report cannot make the requested status transition.")
    run.save()
    ReportRunEvent.objects.create(run=run, actor=actor, action=action, from_status=previous, to_status=run.status, note=note)
    return run


def period_for_schedule(schedule, scheduled_for):
    local_date = timezone.localtime(scheduled_for).date()
    period_end = local_date - timedelta(days=1)
    if schedule.frequency == ReportSchedule.DAILY:
        return period_end, period_end
    if schedule.frequency == ReportSchedule.WEEKLY:
        return period_end - timedelta(days=6), period_end
    if schedule.frequency == ReportSchedule.MONTHLY:
        return period_end.replace(day=1), period_end
    if schedule.frequency == ReportSchedule.QUARTERLY:
        quarter_start_month = ((period_end.month - 1) // 3) * 3 + 1
        return period_end.replace(month=quarter_start_month, day=1), period_end
    return period_end.replace(month=1, day=1), period_end


def advance_schedule(schedule):
    increments = {ReportSchedule.DAILY: relativedelta(days=1), ReportSchedule.WEEKLY: relativedelta(weeks=1), ReportSchedule.MONTHLY: relativedelta(months=1), ReportSchedule.QUARTERLY: relativedelta(months=3), ReportSchedule.ANNUAL: relativedelta(years=1)}
    schedule.next_run_at += increments[schedule.frequency]
    schedule.save(update_fields=("next_run_at", "updated_at"))


def execute_schedule(schedule, scheduled_for=None):
    scheduled_for = scheduled_for or schedule.next_run_at
    period_start, period_end = period_for_schedule(schedule, scheduled_for)
    key = f"schedule:{schedule.pk}:{scheduled_for.isoformat()}:{period_start}:{period_end}"
    with transaction.atomic():
        run, created = ReportRun.objects.get_or_create(
            idempotency_key=key,
            defaults={"definition": schedule.definition, "template_version": schedule.template_version, "schedule": schedule, "output_format": schedule.output_format, "period_start": period_start, "period_end": period_end, "parameters": run_parameters(schedule.definition, schedule.parameters, schedule.template_version), "scheduled_for": scheduled_for, "created_by": schedule.created_by},
        )
    if created or run.status == ReportRun.FAILED:
        generate_report(run)
    if schedule.next_run_at <= scheduled_for:
        advance_schedule(schedule)
    return run, created
