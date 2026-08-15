from __future__ import annotations

import hashlib
import io
from collections import defaultdict
from datetime import datetime
from django.utils import timezone
from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas

from .models import ReportTemplateVersion


class TemplateMappingError(ValueError):
    pass


XLSX_METADATA_NAMES = {
    "GRAND_HEADER": "header",
    "GRAND_TITLE": "title",
    "GRAND_PERIOD": "period",
    "GRAND_PERIOD_START": "period_start",
    "GRAND_PERIOD_END": "period_end",
    "GRAND_CONTROL_ID": "control_id",
    "GRAND_ROW_COUNT": "row_count",
}


def _reference_bytes(template):
    if not template.reference_file:
        raise TemplateMappingError("Upload the department's reference template before mapping it.")
    template.reference_file.open("rb")
    try:
        return template.reference_file.read()
    finally:
        template.reference_file.close()


def _defined_destination(workbook, name, required=False):
    defined_name = workbook.defined_names.get(name)
    if not defined_name:
        if required:
            raise TemplateMappingError(f"The workbook must define the named range {name}.")
        return None
    destinations = list(defined_name.destinations)
    if len(destinations) != 1:
        raise TemplateMappingError(f"{name} must refer to exactly one worksheet range.")
    sheet_name, coordinate = destinations[0]
    if sheet_name not in workbook.sheetnames:
        raise TemplateMappingError(f"{name} points to a worksheet that is not present.")
    return workbook[sheet_name], coordinate


def _reject_external_formulas(workbook):
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=") and "[" in cell.value and "]" in cell.value:
                    raise TemplateMappingError("Mapped workbooks cannot contain formulas linked to external files.")


def inspect_xlsx_template(template, payload=None):
    payload = payload if payload is not None else _reference_bytes(template)
    try:
        workbook = load_workbook(io.BytesIO(payload), data_only=False, keep_links=False)
    except Exception as exc:
        raise TemplateMappingError("The uploaded XLSX workbook could not be opened safely.") from exc
    _reject_external_formulas(workbook)
    sheet, coordinate = _defined_destination(workbook, "GRAND_DATA_AREA", required=True)
    area = sheet[coordinate]
    if not isinstance(area, tuple):
        area = ((area,),)
    elif area and not isinstance(area[0], tuple):
        area = (area,)
    column_count = len(area[0]) if area else 0
    row_capacity = len(area)
    expected_columns = len(template.definition.selected_fields or [])
    if column_count != expected_columns:
        raise TemplateMappingError(f"GRAND_DATA_AREA has {column_count} columns; this report requires {expected_columns}.")
    if row_capacity < 1:
        raise TemplateMappingError("GRAND_DATA_AREA must reserve at least one data row.")
    for name in XLSX_METADATA_NAMES:
        destination = _defined_destination(workbook, name)
        if destination:
            metadata_sheet, metadata_coordinate = destination
            if ":" in metadata_coordinate:
                cells = metadata_sheet[metadata_coordinate]
                flattened = [cell for row in cells for cell in (row if isinstance(row, tuple) else (row,))]
                if len(flattened) != 1:
                    raise TemplateMappingError(f"{name} must refer to one cell.")
    totals = _defined_destination(workbook, "GRAND_TOTALS_AREA")
    if totals:
        totals_sheet, totals_coordinate = totals
        totals_area = totals_sheet[totals_coordinate]
        if not isinstance(totals_area, tuple):
            totals_area = ((totals_area,),)
        elif totals_area and not isinstance(totals_area[0], tuple):
            totals_area = (totals_area,)
        if len(totals_area) != 1 or len(totals_area[0]) != expected_columns:
            raise TemplateMappingError("GRAND_TOTALS_AREA must be one row with the same columns as GRAND_DATA_AREA.")
    return workbook, {
        "mode": ReportTemplateVersion.RENDER_XLSX_TEMPLATE,
        "worksheet": sheet.title,
        "data_area": coordinate,
        "row_capacity": row_capacity,
        "column_count": column_count,
        "metadata_anchors": [name for name in XLSX_METADATA_NAMES if workbook.defined_names.get(name)],
        "has_totals_area": bool(totals),
    }


def inspect_pdf_template(template, payload=None):
    payload = payload if payload is not None else _reference_bytes(template)
    try:
        reader = PdfReader(io.BytesIO(payload))
    except Exception as exc:
        raise TemplateMappingError("The uploaded PDF could not be opened safely.") from exc
    if reader.is_encrypted:
        raise TemplateMappingError("Encrypted PDFs cannot be used as overlay templates.")
    if not reader.pages:
        raise TemplateMappingError("The PDF template has no pages.")
    mappings = list(template.overlay_fields.all())
    if not mappings:
        raise TemplateMappingError("Add at least one reviewed coordinate mapping before validating this PDF.")
    repeating_fields = set()
    for mapping in mappings:
        mapping.full_clean()
        if mapping.page_number > len(reader.pages):
            raise TemplateMappingError(f"{mapping.source_key} points to page {mapping.page_number}, but the PDF has only {len(reader.pages)} page(s).")
        page = reader.pages[mapping.page_number - 1]
        if int(page.get("/Rotate", 0) or 0) % 360:
            raise TemplateMappingError("Rotated PDF pages must be normalized before coordinate mapping.")
        width_mm = float(page.mediabox.width) / 72 * 25.4
        height_mm = float(page.mediabox.height) / 72 * 25.4
        bottom = float(mapping.y_mm) + (float(mapping.row_height_mm) * (mapping.max_rows - 1) if mapping.repeat_for_rows else 0)
        if float(mapping.x_mm) + float(mapping.width_mm) > width_mm or bottom > height_mm:
            raise TemplateMappingError(f"{mapping.source_key} extends beyond page {mapping.page_number}.")
        if mapping.repeat_for_rows:
            repeating_fields.add(mapping.source_key)
    return reader, {
        "mode": ReportTemplateVersion.RENDER_PDF_OVERLAY,
        "page_count": len(reader.pages),
        "mapping_count": len(mappings),
        "repeating_fields": sorted(repeating_fields),
    }


def preflight_template(template, actor):
    if template.render_mode == template.RENDER_NATIVE:
        return {"mode": template.RENDER_NATIVE, "message": "Native GRAND layout uses controlled built-in rendering."}
    payload = _reference_bytes(template)
    checksum = hashlib.sha256(payload).hexdigest()
    if template.render_mode == template.RENDER_XLSX_TEMPLATE:
        _workbook, summary = inspect_xlsx_template(template, payload)
    elif template.render_mode == template.RENDER_PDF_OVERLAY:
        _reader, summary = inspect_pdf_template(template, payload)
    else:
        raise TemplateMappingError("Unsupported template rendering mode.")
    template.mapping_checksum = checksum
    template.mapping_summary = summary
    template.mapping_validated_by = actor
    template.mapping_validated_at = timezone.now()
    template.full_clean()
    template.save(update_fields=("mapping_checksum", "mapping_summary", "mapping_validated_by", "mapping_validated_at"))
    return summary


def _single_named_cell(workbook, name):
    destination = _defined_destination(workbook, name)
    if not destination:
        return None
    sheet, coordinate = destination
    return sheet[coordinate]


def _spreadsheet_value(value):
    if isinstance(value, datetime) and timezone.is_aware(value):
        return timezone.localtime(value).replace(tzinfo=None)
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def generate_mapped_xlsx(run, metadata, rows, totals, selected_fields, display_value):
    payload = _reference_bytes(run.template_version)
    checksum = hashlib.sha256(payload).hexdigest()
    if checksum != run.template_version.mapping_checksum:
        raise TemplateMappingError("The uploaded workbook no longer matches the validated template checksum.")
    workbook, summary = inspect_xlsx_template(run.template_version, payload)
    if len(rows) > summary["row_capacity"]:
        raise TemplateMappingError(f"This workbook reserves {summary['row_capacity']} row(s), but the report produced {len(rows)}. Create a larger template version instead of overwriting the form.")
    metadata_values = {
        "header": metadata["header"], "title": metadata["title"], "period": metadata["period"],
        "period_start": run.period_start, "period_end": run.period_end,
        "control_id": metadata["control_id"], "row_count": len(rows),
    }
    for name, key in XLSX_METADATA_NAMES.items():
        cell = _single_named_cell(workbook, name)
        if cell is not None:
            cell.value = metadata_values[key]
    sheet, coordinate = _defined_destination(workbook, "GRAND_DATA_AREA", required=True)
    area = sheet[coordinate]
    if area and not isinstance(area[0], tuple):
        area = (area,)
    for row_index, cell_row in enumerate(area):
        source = rows[row_index] if row_index < len(rows) else None
        for column_index, cell in enumerate(cell_row):
            cell.value = _spreadsheet_value(source[selected_fields[column_index]]) if source else None
    totals_destination = _defined_destination(workbook, "GRAND_TOTALS_AREA")
    if totals_destination:
        totals_sheet, totals_coordinate = totals_destination
        totals_area = totals_sheet[totals_coordinate]
        if totals_area and not isinstance(totals_area[0], tuple):
            totals_area = (totals_area,)
        for column_index, cell in enumerate(totals_area[0]):
            key = selected_fields[column_index]
            cell.value = totals.get(key) if key in totals else None
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _fit_text(text, font_name, font_size, max_width):
    text = str(text or "")
    size = float(font_size)
    while size > 5 and stringWidth(text, font_name, size) > max_width:
        size -= 0.5
    if stringWidth(text, font_name, size) <= max_width:
        return text, size
    while text and stringWidth(text + "...", font_name, size) > max_width:
        text = text[:-1]
    return text + "...", size


def generate_pdf_overlay(run, metadata, rows, totals, display_value):
    template = run.template_version
    payload = _reference_bytes(template)
    checksum = hashlib.sha256(payload).hexdigest()
    if checksum != template.mapping_checksum:
        raise TemplateMappingError("The uploaded PDF no longer matches the validated template checksum.")
    reader, _summary = inspect_pdf_template(template, payload)
    repeating = list(template.overlay_fields.filter(repeat_for_rows=True))
    if repeating and len(rows) > min(mapping.max_rows for mapping in repeating):
        capacity = min(mapping.max_rows for mapping in repeating)
        raise TemplateMappingError(f"This PDF overlay reserves {capacity} data row(s), but the report produced {len(rows)}. Create a larger template version instead of omitting records.")
    mappings_by_page = defaultdict(list)
    for mapping in template.overlay_fields.all():
        mappings_by_page[mapping.page_number].append(mapping)
    static_values = {
        "header": metadata["header"], "title": metadata["title"], "period": metadata["period"],
        "period_start": run.period_start, "period_end": run.period_end,
        "control_id": metadata["control_id"], "row_count": len(rows),
    }
    static_values.update({f"total:{key}": value for key, value in totals.items()})
    writer = PdfWriter()
    for page_number, page in enumerate(reader.pages, start=1):
        width, height = float(page.mediabox.width), float(page.mediabox.height)
        overlay_buffer = io.BytesIO()
        overlay = canvas.Canvas(overlay_buffer, pagesize=(width, height))
        for mapping in mappings_by_page.get(page_number, []):
            values = []
            if mapping.repeat_for_rows:
                values = [row.get(mapping.source_key) for row in rows[:mapping.max_rows]]
            elif mapping.source_key in static_values:
                values = [static_values[mapping.source_key]]
            else:
                values = [rows[0].get(mapping.source_key) if rows else ""]
            for row_index, value in enumerate(values):
                x = float(mapping.x_mm) / 25.4 * 72
                y_from_top = float(mapping.y_mm) + row_index * float(mapping.row_height_mm)
                y = height - (y_from_top / 25.4 * 72)
                max_width = float(mapping.width_mm) / 25.4 * 72
                text, font_size = _fit_text(display_value(value), "Helvetica", float(mapping.font_size), max_width)
                overlay.setFont("Helvetica", font_size)
                if mapping.alignment == mapping.ALIGN_CENTER:
                    overlay.drawCentredString(x + max_width / 2, y, text)
                elif mapping.alignment == mapping.ALIGN_RIGHT:
                    overlay.drawRightString(x + max_width, y, text)
                else:
                    overlay.drawString(x, y, text)
        overlay.save()
        overlay_buffer.seek(0)
        overlay_page = PdfReader(overlay_buffer).pages[0]
        page.merge_page(overlay_page)
        writer.add_page(page)
    output = io.BytesIO()
    writer.write(output)
    return output.getvalue()
