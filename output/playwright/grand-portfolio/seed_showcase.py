"""Seed a disposable Grand database for browser QA and portfolio screenshots."""

from datetime import timedelta
import io
import os
from pathlib import Path
import sys

import django
from django.utils import timezone
from django.core.files.base import ContentFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName

PROJECT_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(PROJECT_ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "src.settings.dev")
django.setup()

from django.contrib.auth.models import Permission
from assistance.models import AssistanceRequest, AssistanceType
from assistance.services.citizen_service import CitizenProfileService
from departments.models import Department, Plantilla
from home.models import Announcement
from leave_mgt.models import LeaveRequest
from profiles.models import EmployeeProfile
from social_welfare.models import ProgramActivity, SocialWelfareProgram
from reporting.mappers import preflight_template
from reporting.models import ReportRun, ReportSchedule, ReportTemplateVersion
from reporting.presets import seed_mswd_presets
from reporting.services import generate_report, transition_run
from records.models import DepartmentRecord
from records.services import create_record, file_approved_report, transition_record
from users.models import User
from tracepoint.controls import report_discrepancy
from tracepoint.credentials import issue_daily_credential
from tracepoint.handoffs import attach_recipient_code, confirm_handoff, start_scan_session
from tracepoint.models import PacketCheckpoint, PacketDiscrepancy, TrackedPacket
from tracepoint.services import add_checkpoint, add_packet_item, create_packet
from finance.models import (
    FinanceConfigurationItem, FinanceConfigurationRelease, FinanceNumberingSequence,
    FinanceSignatory, FinanceTemplateVersion,
)
from finance.services import preflight_finance_template, transition_release


SHOWCASE_PASSWORD = "GrandShowcase2026!"


def department(name, slug, description, template):
    value, _ = Department.objects.update_or_create(
        slug=slug,
        defaults={
            "name": name,
            "description": description,
            "email": f"{slug}@bocaue.gov.ph",
            "phone": "(044) 123-4567",
            "dashboard_template": template,
        },
    )
    return value


departments = {
    "hr": department(
        "Human Resource Management Office",
        "hr",
        "People, appointments, employee welfare, and workforce services.",
        "home/authed/dashboards/hr.html",
    ),
    "gso": department(
        "General Services Office",
        "gso",
        "Property, supplies, facilities, and procurement support.",
        "home/authed/dashboards/gso.html",
    ),
    "acctg": department(
        "Municipal Accounting Office",
        "acctg",
        "Disbursements, accounting records, compliance, and financial reporting.",
        "home/authed/dashboards/acctg.html",
    ),
    "mpdo": department(
        "Municipal Planning and Development Office",
        "mpdo",
        "Development planning, programs, projects, and municipal data.",
        "",
    ),
    "mswd": department(
        "Municipal Social Welfare and Development Office",
        "mswd",
        "Citizen assistance, social protection, and community support.",
        "home/authed/dashboards/mswd.html",
    ),
}


def employee(username, first_name, last_name, department_key, position, *, staff=False):
    user, _ = User.objects.update_or_create(
        username=username,
        defaults={
            "email": f"{username}@example.gov",
            "first_name": first_name,
            "last_name": last_name,
            "is_active": True,
            "is_staff": staff,
        },
    )
    user.set_password(SHOWCASE_PASSWORD)
    user.save()
    profile = user.employeeprofile
    profile.assigned_department = departments[department_key]
    profile.position_title = position
    profile.employment_type = "REG"
    profile.save()
    return user


hr_user = employee("showcase_hr", "Ana", "Reyes", "hr", "HR Management Officer")
gso_user = employee("showcase_gso", "Marco", "Santos", "gso", "Property Officer")
acctg_user = employee("showcase_acctg", "Elena", "Cruz", "acctg", "Municipal Accountant")
planning_user = employee("showcase_planning", "Paolo", "Mendoza", "mpdo", "Planning Officer")
mswd_user = employee("showcase_mswd", "Liza", "Garcia", "mswd", "Social Welfare Officer")
departments["mswd"].deptHead_or_oic = mswd_user
departments["mswd"].save(update_fields=["deptHead_or_oic"])

tracepoint_permissions = Permission.objects.filter(codename__in=(
    "view_tracepoint_workspace", "prepare_tracked_packets", "print_packet_labels",
    "complete_tracked_packets", "resolve_tracepoint_exceptions", "view_restricted_tracepoint",
))
mswd_user.user_permissions.add(*tracepoint_permissions)
acctg_user.user_permissions.add(*tracepoint_permissions)

for username, first_name, last_name, department_key, position in (
    ("showcase_hr_2", "Mia", "Villanueva", "hr", "Administrative Assistant"),
    ("showcase_hr_3", "Jose", "Navarro", "hr", "Personnel Records Officer"),
    ("showcase_gso_2", "Nico", "Flores", "gso", "Supply Officer"),
    ("showcase_acctg_2", "Cara", "Ramos", "acctg", "Bookkeeper"),
    ("showcase_planning_2", "Iris", "Torres", "mpdo", "Project Development Officer"),
    ("showcase_mswd_2", "Ramon", "Bautista", "mswd", "Community Affairs Officer"),
    ("showcase_mswd_3", "Nina", "Aquino", "mswd", "Social Welfare Assistant"),
):
    employee(username, first_name, last_name, department_key, position)

for title, item_number, grade in (
    ("HR Management Officer", "HR-001", 22),
    ("Personnel Records Officer", "HR-002", 15),
    ("Administrative Assistant", "HR-003", 8),
    ("HR Assistant", "HR-004", 6),
):
    Plantilla.objects.update_or_create(
        item_number=item_number,
        defaults={"title": title, "salary_grade": grade, "department": departments["hr"]},
    )

for profile in EmployeeProfile.objects.filter(assigned_department=departments["hr"]):
    plantilla = Plantilla.objects.filter(
        department=departments["hr"],
        title=profile.position_title,
    ).first()
    if plantilla and profile.plantilla_id != plantilla.id:
        profile.plantilla = plantilla
        profile.save(update_fields=["plantilla"])

today = timezone.localdate()

# Finance showcase data is synthetic and exists only in the disposable portfolio
# database. It demonstrates governance/readiness, never an official LGU rule set.
acctg_approver = User.objects.get(username="showcase_acctg_2")
acctg_user.user_permissions.add(*Permission.objects.filter(content_type__app_label="finance", codename__in=(
    "view_finance_setup", "manage_finance_configuration", "manage_finance_templates",
)))
acctg_approver.user_permissions.add(*Permission.objects.filter(content_type__app_label="finance", codename__in=(
    "view_finance_setup", "approve_finance_configuration",
)))
finance_release, _ = FinanceConfigurationRelease.objects.get_or_create(
    department=departments["acctg"], code=f"synthetic-fy-{today.year}", version=1,
    defaults={
        "title": f"Synthetic Finance Controls FY {today.year}", "fiscal_year": today.year,
        "effective_from": today, "created_by": acctg_user,
    },
)
for category, code, label, configuration in (
    ("transaction_type", "synthetic-disbursement", "Synthetic disbursement type", {"sandbox_only": True}),
    ("document_requirement", "synthetic-dv-checklist", "Synthetic DV checklist", {"items": ["Synthetic obligation reference", "Synthetic supporting schedule"]}),
    ("fund", "synthetic-general-fund", "Synthetic General Fund", {"demonstration": True}),
    ("payment_method", "synthetic-ada", "Synthetic ADA payment method", {"demonstration": True}),
    ("tax_rule", "synthetic-withholding", "Synthetic withholding illustration", {"rate": "demonstration-only", "authoritative": False}),
    ("approval_route", "synthetic-standard-route", "Synthetic Accounting review route", {"steps": ["prepare", "review", "approve"]}),
):
    FinanceConfigurationItem.objects.get_or_create(
        department=departments["acctg"], release=finance_release, category=category, code=code, version=1,
        defaults={"label": label, "configuration": configuration, "effective_from": today, "created_by": acctg_user},
    )
FinanceSignatory.objects.get_or_create(
    department=departments["acctg"], release=finance_release, role_code="approved-by",
    defaults={
        "display_name": "Alex Reyes (Synthetic)", "position_title": "Authorized Official — demonstration",
        "acting": True, "valid_from": today, "created_by": acctg_user,
    },
)
FinanceNumberingSequence.objects.get_or_create(
    department=departments["acctg"], release=finance_release, fiscal_year=today.year, document_type="disbursement-voucher",
    defaults={"prefix": "SYN-DV-", "padding": 6, "next_number": 1, "created_by": acctg_user},
)
finance_template = FinanceTemplateVersion.objects.filter(
    department=departments["acctg"], document_type="disbursement-voucher", version=1,
).first()
if not finance_template:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Synthetic DV"
    sheet.merge_cells("A1:H1")
    sheet["A1"] = "GRAND SYNTHETIC DISBURSEMENT VOUCHER PREVIEW"
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    sheet["A1"].fill = PatternFill("solid", fgColor="17365D")
    sheet["A1"].alignment = Alignment(horizontal="center")
    for name, coordinate in {
        "GRAND_DV_NUMBER": "$G$2", "GRAND_DV_DATE": "$G$3", "GRAND_PAYEE": "$B$4",
        "GRAND_PARTICULARS": "$B$6", "GRAND_GROSS_AMOUNT": "$F$22", "GRAND_TOTAL_DEDUCTIONS": "$G$22",
        "GRAND_NET_AMOUNT": "$H$22", "GRAND_LINE_ITEMS": "$A$12:$D$20", "GRAND_PREPARED_BY": "$B$25",
        "GRAND_CERTIFIED_BY": "$D$25", "GRAND_APPROVED_BY": "$F$25",
    }.items():
        workbook.defined_names.add(DefinedName(name, attr_text=f"'Synthetic DV'!{coordinate}"))
    sheet.print_area = "A1:H28"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    payload = io.BytesIO(); workbook.save(payload)
    finance_template = FinanceTemplateVersion(
        department=departments["acctg"], release=finance_release, document_type="disbursement-voucher",
        version=1, title="Synthetic controlled DV workbook", effective_from=today, created_by=acctg_user,
    )
    finance_template.workbook.save("synthetic-controlled-dv.xlsx", ContentFile(payload.getvalue()), save=False)
    finance_template.full_clean(); finance_template.save()
if finance_template.status == "draft" and not finance_template.preflight_passed:
    preflight_finance_template(finance_template, acctg_user)
finance_release.refresh_from_db()
if finance_release.status == "draft":
    transition_release(finance_release, "submit", acctg_user)
    transition_release(finance_release, "approve", acctg_approver, "Synthetic Accounting approval basis for portfolio demonstration only.")
    transition_release(finance_release, "activate", acctg_approver)

next_monday = today + timedelta(days=(7 - today.weekday()) % 7 or 7)
LeaveRequest.objects.get_or_create(
    employee=hr_user.employeeprofile.leavecredit,
    leave_type="VL",
    start_date=next_monday,
    end_date=next_monday + timedelta(days=1),
    defaults={"status": "PENDING", "notes": "Family appointment"},
)

Announcement.objects.update_or_create(
    slug="grand-digital-services-week",
    defaults={
        "user": hr_user,
        "title": "Digital Services Week",
        "announcement_type": Announcement.INTERNAL,
        "is_pinned": True,
        "published": True,
        "content": "Department teams will review priority digital-service workflows this week.",
    },
)
Announcement.objects.update_or_create(
    slug="grand-records-reminder",
    defaults={
        "user": hr_user,
        "title": "Employee Records Update Reminder",
        "announcement_type": Announcement.INTERNAL,
        "is_pinned": False,
        "published": True,
        "content": "Please verify contact and employment details before month-end.",
    },
)
Announcement.objects.update_or_create(
    slug="grand-public-assistance-advisory",
    defaults={
        "user": mswd_user,
        "title": "Educational Assistance Applications Open",
        "announcement_type": Announcement.PUBLIC,
        "is_pinned": True,
        "published": True,
        "content": "Residents may submit and track educational-assistance requests online.",
    },
)

program, _ = AssistanceType.objects.update_or_create(
    slug="educational-assistance",
    defaults={
        "name": "Educational Assistance",
        "description": "Support for eligible students and families.",
        "requirements": "Birth certificate, certificate of indigency, and current school ID.",
        "is_active": True,
    },
)
for index, (name, email, phone, status, remarks) in enumerate(
    (
        ("Maria Dela Cruz", "maria@example.com", "09170000001", "submitted", "New online application"),
        ("Joshua Reyes", "joshua@example.com", "09170000002", "pending", "Documents ready for initial review"),
        ("Angela Santos", "angela@example.com", "09170000003", "review", "School record verification in progress"),
        ("Carlo Mendoza", "carlo@example.com", "09170000004", "approved", "Approved for scheduled release"),
        ("Maria Dela Cruz", "maria@example.com", "09170000001", "approved", "Prior assistance completed"),
        ("Maria Dela Cruz", "maria@example.com", "09170000001", "denied", "Previous request retained for service history"),
    ),
    start=1,
):
    citizen = CitizenProfileService.get_or_create_citizen(
        full_name=name,
        email=email,
        phone=phone,
    )
    AssistanceRequest.objects.update_or_create(
        reference_code=f"MSWD-SHOWCASE-{index:03d}",
        defaults={
            "assistance_type": program,
            "period": "2026-2027",
            "semester": "1st",
            "full_name": name,
            "email": email,
            "phone": phone,
            "status": status,
            "remarks": remarks,
            "is_active": True,
            "citizen": citizen,
        },
    )
    CitizenProfileService.increment_request_count(citizen)

nutrition_program, _ = SocialWelfareProgram.objects.update_or_create(
    department=departments["mswd"],
    code="MSWD-NUTRITION-2026",
    defaults={
        "name": "Community Nutrition and Family Wellness",
        "program_type": SocialWelfareProgram.TYPE_FEEDING,
        "description": "Coordinated feeding, nutrition education, and family wellness activities in priority barangays.",
        "status": SocialWelfareProgram.STATUS_ACTIVE,
        "coordinator": mswd_user,
        "start_date": today,
        "end_date": today + timedelta(days=180),
        "created_by": mswd_user,
        "updated_by": mswd_user,
    },
)
family_program, _ = SocialWelfareProgram.objects.update_or_create(
    department=departments["mswd"],
    code="MSWD-FAMILY-2026",
    defaults={
        "name": "Family Development and Protection",
        "program_type": SocialWelfareProgram.TYPE_SEMINAR,
        "description": "Orientations, referral pathways, and community sessions that strengthen family support systems.",
        "status": SocialWelfareProgram.STATUS_ACTIVE,
        "coordinator": mswd_user,
        "start_date": today - timedelta(days=45),
        "end_date": today + timedelta(days=120),
        "created_by": mswd_user,
        "updated_by": mswd_user,
    },
)
for program_record, title, activity_type, start_offset, venue, status, expected, actual, outcome in (
    (
        nutrition_program,
        "Community Feeding and Nutrition Seminar",
        ProgramActivity.TYPE_FEEDING,
        7,
        "Barangay Mabuhay Multi-Purpose Hall",
        ProgramActivity.STATUS_PLANNED,
        120,
        None,
        "",
    ),
    (
        family_program,
        "Parent Effectiveness and Child Protection Orientation",
        ProgramActivity.TYPE_ORIENTATION,
        14,
        "Municipal Training Hall",
        ProgramActivity.STATUS_PLANNED,
        85,
        None,
        "",
    ),
    (
        nutrition_program,
        "Nutrition Screening and Family Referral Day",
        ProgramActivity.TYPE_OUTREACH,
        -7,
        "MSWD Community Center",
        ProgramActivity.STATUS_COMPLETED,
        75,
        68,
        "Families received nutrition screening results and service referral materials.",
    ),
):
    starts_at = timezone.now() + timedelta(days=start_offset)
    ProgramActivity.objects.update_or_create(
        program=program_record,
        title=title,
        defaults={
            "activity_type": activity_type,
            "starts_at": starts_at,
            "ends_at": starts_at + timedelta(hours=3),
            "venue": venue,
            "status": status,
            "expected_attendance": expected,
            "actual_attendance": actual,
            "outcome_notes": outcome,
            "created_by": mswd_user,
            "updated_by": mswd_user,
        },
    )

report_presets = seed_mswd_presets(mswd_user)
report_definitions = {definition.slug: definition for definition, _ in report_presets}
validated_template = report_definitions["assistance-volume-status"].current_template
validated_template.fidelity_status = validated_template.OFFICIAL
validated_template.fidelity_notes = "Synthetic side-by-side comparison completed against the showcase MSWD monthly form."
validated_template.fidelity_validated_by = mswd_user
validated_template.fidelity_validated_at = timezone.now()
validated_template.save(update_fields=("fidelity_status", "fidelity_notes", "fidelity_validated_by", "fidelity_validated_at"))

# A synthetic current-office workbook demonstrates controlled compatibility without
# committing citizen data or a department-owned source file to the repository.
mapped_template = ReportTemplateVersion.objects.filter(
    definition=report_definitions["assistance-volume-status"], version=2,
).first()
if not mapped_template:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Monthly Assistance Form"
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "MUNICIPAL SOCIAL WELFARE AND DEVELOPMENT OFFICE"
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    sheet["A1"].fill = PatternFill("solid", fgColor="17365D")
    sheet["A1"].alignment = Alignment(horizontal="center")
    for cell, label in zip(("A6", "B6", "C6"), ("Assistance type", "Status", "Requests")):
        sheet[cell] = label
        sheet[cell].font = Font(bold=True, color="FFFFFF")
        sheet[cell].fill = PatternFill("solid", fgColor="2B579A")
    thin = Side(style="thin", color="9AA7B2")
    for row in sheet.iter_rows(min_row=6, max_row=18, min_col=1, max_col=3):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 14
    workbook.defined_names.add(DefinedName("GRAND_TITLE", attr_text="'Monthly Assistance Form'!$A$2"))
    workbook.defined_names.add(DefinedName("GRAND_PERIOD", attr_text="'Monthly Assistance Form'!$A$3"))
    workbook.defined_names.add(DefinedName("GRAND_CONTROL_ID", attr_text="'Monthly Assistance Form'!$A$4"))
    workbook.defined_names.add(DefinedName("GRAND_DATA_AREA", attr_text="'Monthly Assistance Form'!$A$7:$C$17"))
    workbook.defined_names.add(DefinedName("GRAND_TOTALS_AREA", attr_text="'Monthly Assistance Form'!$A$18:$C$18"))
    workbook_bytes = io.BytesIO()
    workbook.save(workbook_bytes)
    mapped_template = ReportTemplateVersion(
        definition=report_definitions["assistance-volume-status"], version=2,
        title="Current MSWD Monthly Assistance Form", render_mode=ReportTemplateVersion.RENDER_XLSX_TEMPLATE,
        reference_kind=ReportTemplateVersion.REFERENCE_XLSX,
        mapping_notes="Synthetic workbook with controlled named ranges for portfolio demonstration.", created_by=mswd_user,
    )
    mapped_template.reference_file.save("synthetic-mswd-monthly-form.xlsx", ContentFile(workbook_bytes.getvalue()), save=False)
    mapped_template.full_clean()
    mapped_template.save()
preflight_template(mapped_template, mswd_user)
showcase_period_start = today.replace(day=1)
showcase_period_end = today
for slug, desired_status, output_format in (
    ("assistance-volume-status", ReportRun.APPROVED, "pdf"),
    ("program-activity-accomplishment", ReportRun.REVIEWED, "xlsx"),
    ("department-workload", ReportRun.GENERATED, "csv"),
):
    definition = report_definitions[slug]
    run, _ = ReportRun.objects.get_or_create(
        idempotency_key=f"showcase:{slug}:{showcase_period_start}:{showcase_period_end}",
        defaults={
            "definition": definition,
            "template_version": definition.current_template,
            "output_format": output_format,
            "period_start": showcase_period_start,
            "period_end": showcase_period_end,
            "parameters": {},
            "created_by": mswd_user,
        },
    )
    if run.status in (ReportRun.DRAFT, ReportRun.FAILED):
        generate_report(run)
    if desired_status in (ReportRun.REVIEWED, ReportRun.APPROVED) and run.status == ReportRun.GENERATED:
        transition_run(run, "review", mswd_user, "Synthetic showcase figures reviewed.")
    if desired_status == ReportRun.APPROVED and run.status == ReportRun.REVIEWED:
        transition_run(run, "approve", mswd_user, "Synthetic showcase output approved.")

approved_run = ReportRun.objects.get(
    idempotency_key=f"showcase:assistance-volume-status:{showcase_period_start}:{showcase_period_end}"
)
file_approved_report(approved_run, mswd_user)

program_record = DepartmentRecord.objects.filter(
    department=departments["mswd"], title="Community Nutrition Program Accomplishment File"
).first()
if not program_record:
    program_record = create_record(
        department=departments["mswd"], actor=mswd_user,
        title="Community Nutrition Program Accomplishment File",
        description="Approved program file linking the operating program with its signed accomplishment note.",
        classification=DepartmentRecord.CLASS_PROGRAM, custodian=mswd_user, retention_years=5,
        retention_notes="Synthetic five-year departmental retention example for portfolio QA.",
        sources=(nutrition_program,),
        uploaded_file=ContentFile(b"Synthetic signed accomplishment note.", name="nutrition-accomplishment-note.txt"),
        uploaded_description="Synthetic signed accomplishment note for visual testing.",
    )
    transition_record(program_record, "submit", mswd_user, "Prepared for records review.")
    transition_record(program_record, "approve", mswd_user, "Source and file integrity checked.")

request_for_record = AssistanceRequest.objects.get(reference_code="MSWD-SHOWCASE-002")
if not DepartmentRecord.objects.filter(
    department=departments["mswd"], title="Assistance Case File — MSWD-SHOWCASE-002"
).exists():
    assistance_record = create_record(
        department=departments["mswd"], actor=mswd_user,
        title="Assistance Case File — MSWD-SHOWCASE-002",
        description="Restricted case record linked to its live Assistance request.",
        classification=DepartmentRecord.CLASS_ASSISTANCE,
        confidentiality=DepartmentRecord.CONFIDENTIALITY_CONFIDENTIAL,
        custodian=mswd_user, sources=(request_for_record,),
    )
    transition_record(assistance_record, "submit", mswd_user, "Ready for records review.")

if not DepartmentRecord.objects.filter(
    department=departments["mswd"], title="2026 Records Retention Schedule Review"
).exists():
    create_record(
        department=departments["mswd"], actor=mswd_user,
        title="2026 Records Retention Schedule Review",
        description="Draft working file for aligning department records with the approved retention schedule.",
        classification=DepartmentRecord.CLASS_GENERAL, custodian=mswd_user,
    )

schedule_definition = report_definitions["assistance-volume-status"]
ReportSchedule.objects.update_or_create(
    definition=schedule_definition,
    name="Monthly assistance status report",
    defaults={
        "template_version": schedule_definition.current_template,
        "frequency": ReportSchedule.MONTHLY,
        "output_format": "pdf",
        "next_run_at": timezone.now() + timedelta(days=5),
        "parameters": {},
        "is_active": True,
        "created_by": mswd_user,
    },
)

# Synthetic custody routes demonstrate physical-paper traceability without using
# production employees, reports, records, or citizen information.
route_packet = TrackedPacket.objects.filter(title="Monthly Assistance Voucher Bundle — July 2026").first()
if not route_packet:
    route_packet = create_packet(
        actor=mswd_user,
        title="Monthly Assistance Voucher Bundle — July 2026",
        contents_manifest="One approved summary report, twelve voucher folders, and signed supporting attachments.",
        expected_document_count=13,
        expected_page_count=86,
        confidentiality=TrackedPacket.RESTRICTED,
        final_destination_department=departments["acctg"],
        final_destination_employee=acctg_user,
        report_run=approved_run,
    )

if route_packet.status == TrackedPacket.DRAFT and not route_packet.voucher_items.exists():
    for number, title in enumerate((
        "Emergency assistance voucher — Garcia household",
        "Medical assistance voucher — Reyes household",
        "Burial assistance voucher — Santos household",
        "Transportation assistance voucher — Mendoza household",
    ), start=1):
        add_packet_item(
            packet=route_packet,
            actor=mswd_user,
            title=title,
            description=f"Synthetic showcase voucher {number}; no real citizen data.",
            expected_attachment_count=4,
            expected_page_count=7,
        )

if route_packet.status == TrackedPacket.DRAFT:
    if not route_packet.checkpoints.filter(label="Accounting signature and initial verification").exists():
        add_checkpoint(
            packet=route_packet,
            actor=mswd_user,
            department=departments["acctg"],
            employee=acctg_user,
            purpose=PacketCheckpoint.SIGNATURE,
            label="Accounting signature and initial verification",
            instructions="Verify voucher totals, sign the control sheet, then route onward.",
        )
    if not route_packet.checkpoints.filter(label="Planning certification").exists():
        add_checkpoint(
            packet=route_packet,
            actor=mswd_user,
            department=departments["mpdo"],
            employee=planning_user,
            purpose=PacketCheckpoint.CERTIFICATION,
            label="Planning certification",
            instructions="Confirm the linked accomplishment figures before final return.",
        )
    if not route_packet.checkpoints.filter(label="Final Accounting custody").exists():
        add_checkpoint(
            packet=route_packet,
            actor=mswd_user,
            department=departments["acctg"],
            employee=acctg_user,
            purpose=PacketCheckpoint.RELEASE,
            label="Final Accounting custody",
            instructions="Retain the completed voucher bundle as the declared terminal destination.",
        )


def confirm_showcase_receipt(packet, receiver, operator, key, checkpoint=None, terminal=False):
    issued = issue_daily_credential(employee=receiver, actor=receiver, replace=True)
    scan = start_scan_session(packet=packet, operator=operator, idempotency_key=key)
    attach_recipient_code(session=scan, operator=operator, token=issued.token)
    return confirm_handoff(
        session=scan,
        operator=operator,
        receipt_note="Physical contents counted and accepted for the next processing step.",
        checkpoint=checkpoint,
        terminal_delivery=terminal,
    )


if route_packet.status == TrackedPacket.DRAFT:
    confirm_showcase_receipt(route_packet, mswd_user, mswd_user, "showcase-tracepoint-activate")
    route_packet.refresh_from_db()
if route_packet.status == TrackedPacket.ACTIVE and route_packet.current_holder_id == mswd_user.pk:
    checkpoint = route_packet.checkpoints.filter(label="Accounting signature and initial verification").first()
    confirm_showcase_receipt(
        route_packet, acctg_user, acctg_user, "showcase-tracepoint-accounting-signature", checkpoint=checkpoint,
    )
    route_packet.refresh_from_db()
if route_packet.status == TrackedPacket.ACTIVE and route_packet.current_holder_id == acctg_user.pk:
    checkpoint = route_packet.checkpoints.filter(label="Planning certification").first()
    confirm_showcase_receipt(
        route_packet, planning_user, planning_user, "showcase-tracepoint-planning", checkpoint=checkpoint,
    )
    route_packet.refresh_from_db()
if route_packet.status == TrackedPacket.ACTIVE and route_packet.current_holder_id == planning_user.pk:
    checkpoint = route_packet.checkpoints.filter(label="Final Accounting custody").first()
    final_handoff = confirm_showcase_receipt(
        route_packet,
        acctg_user,
        acctg_user,
        "showcase-tracepoint-accounting-terminal",
        checkpoint=checkpoint,
        terminal=True,
    )
    route_packet.refresh_from_db()
    if not route_packet.discrepancies.exists():
        report_discrepancy(
            packet=route_packet,
            actor=acctg_user,
            category=PacketDiscrepancy.MISSING_CONTENTS,
            description="Control sheet notes thirteen items; receiving count found twelve pending preparer verification.",
            related_handoff=final_handoff,
        )

active_packet = TrackedPacket.objects.filter(title="Nutrition Program Liquidation Packet").first()
if not active_packet:
    active_packet = create_packet(
        actor=mswd_user,
        title="Nutrition Program Liquidation Packet",
        contents_manifest="Activity accomplishment report, attendance totals, receipts, and liquidation worksheet.",
        expected_document_count=8,
        confidentiality=TrackedPacket.INTERNAL,
        final_destination_department=departments["acctg"],
        department_record=program_record,
    )
if active_packet.status == TrackedPacket.DRAFT:
    confirm_showcase_receipt(active_packet, mswd_user, mswd_user, "showcase-tracepoint-active")

draft_packet = TrackedPacket.objects.filter(title="Family Development Seminar Records").first()
if not draft_packet:
    draft_packet = create_packet(
        actor=mswd_user,
        title="Family Development Seminar Records",
        contents_manifest="Seminar completion note, aggregate attendance sheet, and supporting program documents.",
        expected_document_count=5,
        confidentiality=TrackedPacket.INTERNAL,
        final_destination_department=departments["mswd"],
    )

print("Showcase database seeded.")
print(f"Dashboard users: showcase_hr, showcase_gso, showcase_acctg, showcase_planning, showcase_mswd")
print(f"Password: {SHOWCASE_PASSWORD}")
print(f"Records detail: {program_record.get_absolute_url()}")
print(f"TracePoint route detail: {route_packet.get_absolute_url()}")
print(f"Finance setup detail: /finance/setup/releases/{finance_release.pk}/")
