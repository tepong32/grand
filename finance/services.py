from __future__ import annotations

import hashlib
import io
import json
import zipfile
from datetime import date

from django.core.exceptions import PermissionDenied, ValidationError
from django.db import transaction
from django.forms.models import model_to_dict
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PageMargins

from .access import can_approve_finance_configuration, can_manage_finance_configuration, can_manage_finance_templates
from .models import (
    FinanceAuditEvent, FinanceConfigurationItem, FinanceConfigurationRelease,
    FinanceNumberingSequence, FinanceParty, FinancePostingRule, FinancePostingRuleLine,
    FinanceSignatory, FinanceTemplateVersion, FinanceTransactionVariant,
)


class FinanceTemplateError(ValueError):
    pass


def build_finance_starter_workbook(values):
    """Build a deliberately plain, editable, macro-free DV starter for local comparison."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DV Starter"
    instructions = workbook.create_sheet("Read Me First")

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    plain_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="E7E6E6")
    starter_fill = PatternFill("solid", fgColor="FFF2CC")

    sheet.merge_cells("A1:F1")
    sheet["A1"] = values["lgu_name"]
    sheet["A1"].font = Font(name="Arial", size=12, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A2:F2")
    sheet["A2"] = values["finance_office_name"]
    sheet["A2"].font = Font(name="Arial", size=10, bold=True)
    sheet["A2"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A3:F3")
    sheet["A3"] = values["form_title"]
    sheet["A3"].font = Font(name="Arial", size=14, bold=True)
    sheet["A3"].alignment = Alignment(horizontal="center")

    sheet.merge_cells("A4:D4")
    sheet["A4"] = values["form_reference"]
    sheet["A4"].fill = starter_fill
    sheet["A4"].font = Font(name="Arial", size=9, italic=True)
    sheet["E4"], sheet["F4"] = "DV No.", ""
    sheet["E5"], sheet["F5"] = "DV Date", ""
    sheet["A5"], sheet["B5"] = "Fund", ""
    sheet.merge_cells("B5:D5")
    sheet["A6"], sheet["B6"] = "Payee", ""
    sheet.merge_cells("B6:F6")
    sheet["A7"], sheet["B7"] = "Address", ""
    sheet.merge_cells("B7:D7")
    sheet["E7"], sheet["F7"] = "TIN", ""
    sheet["A8"], sheet["B8"] = "OBR / ORS / ALOBS reference", ""
    sheet.merge_cells("B8:D8")
    sheet["E8"], sheet["F8"] = "Claim reference", ""
    sheet["A9"], sheet["B9"], sheet["C9"] = "Particulars", "Account code", "Amount"
    sheet.merge_cells("C9:F9")

    detail_start = 10
    detail_end = detail_start + int(values["particulars_rows"]) - 1
    for row in range(detail_start, detail_end + 1):
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        sheet.row_dimensions[row].height = 22
    totals_row = detail_end + 1
    for offset, (label, name) in enumerate((
        ("Gross amount", "GRAND_GROSS_AMOUNT"),
        ("Total deductions", "GRAND_TOTAL_DEDUCTIONS"),
        ("Net amount", "GRAND_NET_AMOUNT"),
    )):
        row = totals_row + offset
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        sheet.cell(row, 1).value = label
        sheet.cell(row, 1).alignment = Alignment(horizontal="right")
        sheet.cell(row, 5).value = "PHP"
        sheet.cell(row, 6).number_format = '#,##0.00'
        workbook.defined_names.add(DefinedName(name, attr_text=f"'DV Starter'!$F${row}"))

    certification_row = totals_row + 4
    certification_labels = (
        ("A", "B", values["prepared_label"], "GRAND_PREPARED_BY"),
        ("C", "D", values["certified_label"], "GRAND_CERTIFIED_BY"),
        ("E", "F", values["approved_label"], "GRAND_APPROVED_BY"),
    )
    for start_col, end_col, label, name in certification_labels:
        sheet.merge_cells(f"{start_col}{certification_row}:{end_col}{certification_row}")
        sheet[f"{start_col}{certification_row}"] = label
        sheet[f"{start_col}{certification_row}"].fill = header_fill
        sheet[f"{start_col}{certification_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet.merge_cells(f"{start_col}{certification_row + 1}:{end_col}{certification_row + 3}")
        sheet[f"{start_col}{certification_row + 1}"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
        workbook.defined_names.add(DefinedName(name, attr_text=f"'DV Starter'!${start_col}${certification_row + 1}"))

    receipt_row = certification_row + 4
    sheet.merge_cells(start_row=receipt_row, start_column=1, end_row=receipt_row, end_column=6)
    sheet.cell(receipt_row, 1).value = "Payment / receipt details — reserved for the locally accepted Treasury process"
    sheet.cell(receipt_row, 1).fill = header_fill
    sheet.cell(receipt_row, 1).font = Font(name="Arial", size=9, bold=True)
    sheet.merge_cells(start_row=receipt_row + 1, start_column=1, end_row=receipt_row + 2, end_column=6)
    sheet.cell(receipt_row + 1, 1).value = "Do not add check, claimant, or receipt fields here until the F8 route and local form are confirmed."
    sheet.cell(receipt_row + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")

    footer_row = receipt_row + 3
    sheet.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=6)
    sheet.cell(footer_row, 1).value = values.get("footer_note", "")
    sheet.cell(footer_row, 1).font = Font(name="Arial", size=8, italic=True)
    sheet.cell(footer_row, 1).alignment = Alignment(horizontal="center", wrap_text=True)

    fixed_names = {
        "GRAND_DV_NUMBER": "$F$4",
        "GRAND_DV_DATE": "$F$5",
        "GRAND_PAYEE": "$B$6",
        "GRAND_PARTICULARS": "$A$10",
        "GRAND_LINE_ITEMS": f"$A${detail_start}:$C${detail_end}",
    }
    for name, coordinate in fixed_names.items():
        workbook.defined_names.add(DefinedName(name, attr_text=f"'DV Starter'!{coordinate}"))

    for row in sheet.iter_rows(min_row=4, max_row=receipt_row + 2, min_col=1, max_col=6):
        for cell in row:
            cell.border = plain_border
            cell.font = cell.font.copy(name="Arial", size=cell.font.sz or 9)
            cell.alignment = cell.alignment.copy(vertical=cell.alignment.vertical or "center", wrap_text=True)
    for cell in sheet[9]:
        cell.fill = header_fill
        cell.font = Font(name="Arial", size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[9].height = 28
    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 8
    sheet.column_dimensions["D"].width = 8
    sheet.column_dimensions["E"].width = 12
    sheet.column_dimensions["F"].width = 18
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A9"
    sheet.print_area = f"A1:F{footer_row}"
    sheet.print_title_rows = "1:9"
    sheet.page_setup.paperSize = {"a4": "9", "letter": "1", "legal": "5"}[values["paper_size"]]
    sheet.page_setup.orientation = values["orientation"]
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4, header=0.15, footer=0.15)

    instructions.column_dimensions["A"].width = 28
    instructions.column_dimensions["B"].width = 90
    instructions["A1"], instructions["B1"] = "Editable DV starter", "What to do"
    instructions["A1"].font = instructions["B1"].font = Font(name="Arial", size=11, bold=True)
    guidance = (
        ("Status", "This workbook is a starter for local review. It is not automatically an official COA, DBM, or municipal form."),
        ("Safe edits", "Change ordinary labels, wording, row heights, borders, logo space, paper size, and signatory captions in Excel."),
        ("Keep these controls", "Do not delete workbook-level names beginning GRAND_. GRAND uses them to place reviewed data without macros."),
        ("Local comparison", "Compare the starter side by side with the current blank form, a redacted completed sample, the actual signature route, and printer output."),
        ("Upload", "Upload the edited .xlsx as a new draft Finance Template version, record its authority and comparison references, then run preflight."),
        ("Copies", f"Starter assumption: {values['default_copy_count']} copy/copies. Confirm this locally before activation."),
        ("Security", "Do not place real citizen, supplier, employee, bank, tax, or claim data in the reusable blank workbook."),
    )
    for row_index, (topic, explanation) in enumerate(guidance, start=2):
        instructions.cell(row_index, 1).value = topic
        instructions.cell(row_index, 1).font = Font(name="Arial", size=10, bold=True)
        instructions.cell(row_index, 2).value = explanation
        instructions.cell(row_index, 2).alignment = Alignment(wrap_text=True, vertical="top")
        instructions.row_dimensions[row_index].height = 34
    instructions.sheet_view.showGridLines = False

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _snapshot(instance):
    data = model_to_dict(instance)
    for key, value in tuple(data.items()):
        if hasattr(value, "isoformat"):
            data[key] = value.isoformat()
        elif isinstance(value, uuid_types()):
            data[key] = str(value)
        elif hasattr(value, "name"):
            data[key] = value.name
    return data


def uuid_types():
    import uuid
    return (uuid.UUID,)


def record_event(instance, actor, action, reason="", evidence=None):
    release = instance if isinstance(instance, FinanceConfigurationRelease) else getattr(instance, "release", None)
    snapshot = _snapshot(instance)
    if evidence:
        snapshot["workflow_exemption"] = evidence
    return FinanceAuditEvent.objects.create(
        department=instance.department, release=release, target_type=instance._meta.model_name,
        target_id=str(instance.pk), action=action, actor=actor, reason=reason,
        snapshot=snapshot,
    )


def posting_rule_snapshot(rule):
    """Return a stable, plain-data posting recipe suitable for an immutable transaction snapshot."""
    lines = list(rule.lines.order_by("sequence", "pk"))
    if rule.accounting_effect == FinancePostingRule.JOURNAL_ENTRY:
        if not lines:
            raise ValidationError("A journal-producing posting rule needs debit and credit instructions.")
        if not {line.side for line in lines}.issuperset({FinancePostingRuleLine.DEBIT, FinancePostingRuleLine.CREDIT}):
            raise ValidationError("The selected posting rule must contain at least one debit and one credit instruction.")
    elif lines:
        raise ValidationError("An explicit no-entry rule cannot contain debit or credit instructions.")
    snapshot = {
        "schema_version": 1,
        "rule_public_id": str(rule.public_id),
        "rule_code": rule.code,
        "variant_public_id": str(rule.variant.public_id),
        "variant_code": rule.variant.code,
        "event_kind": rule.event_kind,
        "event_label": rule.get_event_kind_display(),
        "recognition_point": rule.recognition_point,
        "recognition_point_label": rule.get_recognition_point_display(),
        "accounting_effect": rule.accounting_effect,
        "accounting_effect_label": rule.get_accounting_effect_display(),
        "title": rule.title,
        "description": rule.description,
        "authority_reference": rule.authority_reference,
        "release_id": rule.variant.release_id,
        "release_code": rule.variant.release.code,
        "release_version": rule.variant.release.version,
        "lines": [
            {
                "sequence": line.sequence,
                "label": line.label,
                "side": line.side,
                "account_source": line.account_source,
                "amount_source": line.amount_source,
                "mapping_code": line.mapping_code,
                "ledger_account_code": line.ledger_account_code,
                "memo": line.memo,
            }
            for line in lines
        ],
    }
    encoded = json.dumps(snapshot, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return snapshot, hashlib.sha256(encoded).hexdigest()


def payment_event_policy_error(rules):
    """Reject an issuance-time settlement policy that would not undo/reapply its ledger effect."""
    by_event = {rule.event_kind: rule for rule in rules}
    payment = by_event.get(FinancePostingRule.PAYMENT)
    if not payment or payment.accounting_effect != FinancePostingRule.JOURNAL_ENTRY:
        return ""
    if payment.recognition_point != FinancePostingRule.PAYMENT_ISSUANCE:
        return ""
    missing_effects = [
        kind for kind in (FinancePostingRule.CANCELLATION, FinancePostingRule.REPLACEMENT)
        if not by_event.get(kind) or by_event[kind].accounting_effect != FinancePostingRule.JOURNAL_ENTRY
    ]
    if not missing_effects:
        return ""
    labels = dict(FinancePostingRule.EVENT_KIND_CHOICES)
    return (
        "A payment JEV recorded at instrument issuance requires journal-producing "
        + " and ".join(labels[kind].lower() for kind in missing_effects)
        + " rules so a cancelled or replacement instrument cannot leave an unexplained ledger effect."
    )


@transaction.atomic
def create_recognition_posting_starter(variant, actor):
    """Create an editable three-part payable-recognition recipe inside one draft release."""
    if not can_manage_finance_configuration(actor, variant.department):
        raise PermissionDenied
    variant = FinanceTransactionVariant.objects.select_for_update().select_related("release", "department").get(pk=variant.pk)
    if variant.release.status != "draft" or variant.status != "draft":
        raise ValidationError("A posting starter can be added only to a draft transaction variant.")
    if variant.posting_rules.filter(event_kind=FinancePostingRule.RECOGNITION).exists():
        raise ValidationError("This transaction variant already has a recognition posting rule.")
    rule = FinancePostingRule(
        variant=variant,
        code=f"{variant.code}-recognition",
        title=f"Recognize {variant.label}",
        event_kind=FinancePostingRule.RECOGNITION,
        recognition_point=FinancePostingRule.DV_VALIDATION,
        description=(
            "At Accounting validation, debit each reviewed voucher allocation, credit each configured deduction, "
            "and credit the remaining net amount to the transaction's configured payable account."
        ),
        authority_reference=(
            "EDIT BEFORE SUBMISSION — compare with the locally accepted accounting entry, current COA guidance, "
            "and the Accounting office's reviewed recognition decision."
        ),
        created_by=actor,
    )
    rule.full_clean()
    rule.save()
    FinancePostingRuleLine.objects.bulk_create((
        FinancePostingRuleLine(
            rule=rule, sequence=10, label="Debit each reviewed allocation account",
            side=FinancePostingRuleLine.DEBIT,
            account_source=FinancePostingRuleLine.ALLOCATION_ACCOUNTS,
            amount_source=FinancePostingRuleLine.EACH_ALLOCATION,
            memo="Reviewed voucher allocation",
        ),
        FinancePostingRuleLine(
            rule=rule, sequence=20, label="Credit each deduction payable",
            side=FinancePostingRuleLine.CREDIT,
            account_source=FinancePostingRuleLine.DEDUCTION_MAPPINGS,
            amount_source=FinancePostingRuleLine.EACH_DEDUCTION,
            memo="Voucher deduction / withholding",
        ),
        FinancePostingRuleLine(
            rule=rule, sequence=30, label="Credit the net transaction payable",
            side=FinancePostingRuleLine.CREDIT,
            account_source=FinancePostingRuleLine.PAYABLE_MAPPING,
            amount_source=FinancePostingRuleLine.NET,
            memo="Net payable",
        ),
    ))
    record_event(rule, actor, "posting_rule_starter_created")
    return rule


@transaction.atomic
def create_payment_event_posting_starters(variant, actor):
    """Create editable payment-cycle starters, including explicit no-entry decisions."""
    if not can_manage_finance_configuration(actor, variant.department):
        raise PermissionDenied
    variant = FinanceTransactionVariant.objects.select_for_update().select_related(
        "release", "department",
    ).get(pk=variant.pk)
    if variant.release.status != "draft" or variant.status != "draft":
        raise ValidationError("Payment-event starters can be added only to a draft transaction variant.")
    definitions = (
        (
            FinancePostingRule.PAYMENT,
            FinancePostingRule.PAYMENT_RELEASE,
            FinancePostingRule.JOURNAL_ENTRY,
            "Record payment on actual release",
            "At actual release, debit the transaction payable and credit the configured bank/cash account for this instrument.",
            (
                (10, "Debit the released payable", FinancePostingRuleLine.DEBIT, FinancePostingRuleLine.PAYABLE_MAPPING),
                (20, "Credit the releasing bank/cash account", FinancePostingRuleLine.CREDIT, FinancePostingRuleLine.BANK_MAPPING),
            ),
        ),
        (
            FinancePostingRule.CANCELLATION,
            FinancePostingRule.PAYMENT_CANCELLATION,
            FinancePostingRule.NO_ENTRY,
            "Record pre-release check cancellation",
            "Record the cancelled or spoiled instrument and its reason without a JEV when no payment was recognized before release.",
            (),
        ),
        (
            FinancePostingRule.REVERSAL,
            FinancePostingRule.PAYMENT_RETURN,
            FinancePostingRule.JOURNAL_ENTRY,
            "Restore a bank-returned payment",
            "After bank-return evidence and independent Accounting review, debit the configured bank/cash account and credit the transaction payable to reverse the earlier release entry.",
            (
                (10, "Debit the bank/cash account restored by the return", FinancePostingRuleLine.DEBIT, FinancePostingRuleLine.BANK_MAPPING),
                (20, "Credit the payable restored by the return", FinancePostingRuleLine.CREDIT, FinancePostingRuleLine.PAYABLE_MAPPING),
            ),
        ),
        (
            FinancePostingRule.REPLACEMENT,
            FinancePostingRule.PAYMENT_REPLACEMENT,
            FinancePostingRule.NO_ENTRY,
            "Record pre-release replacement check",
            "Record the replacement lineage without a JEV when the related cancelled check never reached payment recognition.",
            (),
        ),
        (
            FinancePostingRule.REMITTANCE,
            FinancePostingRule.DEDUCTION_REMITTANCE,
            FinancePostingRule.JOURNAL_ENTRY,
            "Record deduction or withholding remittance",
            "At accepted remittance, debit each remitted deduction payable and credit the configured bank/cash account.",
            (
                (10, "Debit each remitted deduction payable", FinancePostingRuleLine.DEBIT, FinancePostingRuleLine.DEDUCTION_MAPPINGS),
                (20, "Credit the remitting bank/cash account", FinancePostingRuleLine.CREDIT, FinancePostingRuleLine.BANK_MAPPING),
            ),
        ),
    )
    created = []
    for event_kind, point, effect, title, description, lines in definitions:
        if variant.posting_rules.filter(event_kind=event_kind).exists():
            continue
        rule = FinancePostingRule(
            variant=variant,
            code=f"{variant.code}-{event_kind}",
            title=title,
            event_kind=event_kind,
            recognition_point=point,
            accounting_effect=effect,
            description=description,
            authority_reference=(
                "EDIT BEFORE SUBMISSION — compare this starter with the locally accepted payment, cancellation, "
                "replacement, returned-item, or remittance treatment and the current COA/local accounting basis."
            ),
            created_by=actor,
        )
        rule.full_clean()
        rule.save()
        for sequence, label, side, account_source in lines:
            amount_source = (
                FinancePostingRuleLine.EACH_DEDUCTION
                if account_source == FinancePostingRuleLine.DEDUCTION_MAPPINGS
                else FinancePostingRuleLine.TOTAL_DEDUCTIONS
                if event_kind == FinancePostingRule.REMITTANCE and account_source == FinancePostingRuleLine.BANK_MAPPING
                else FinancePostingRuleLine.EVENT_AMOUNT
            )
            FinancePostingRuleLine.objects.create(
                rule=rule,
                sequence=sequence,
                label=label,
                side=side,
                account_source=account_source,
                amount_source=amount_source,
                memo=label,
            )
        record_event(rule, actor, "payment_event_posting_starter_created")
        created.append(rule)
    if not created:
        raise ValidationError("This transaction variant already has payment-cycle rules for every starter event.")
    return created


@transaction.atomic
def transition_release(release, action, actor, reason=""):
    release = FinanceConfigurationRelease.objects.select_for_update().get(pk=release.pk)
    workflow_exemption = None
    if action == "submit":
        if not can_manage_finance_configuration(actor, release.department):
            raise PermissionDenied
        if release.status != "draft":
            raise ValidationError("Only draft releases can be submitted.")
        for variant in release.transaction_variants.all():
            if not variant.posting_rules.exists():
                raise ValidationError(f"{variant.label} needs at least one locally reviewed posting rule.")
            event_kinds = set(variant.posting_rules.values_list("event_kind", flat=True))
            required_payment_events = {
                FinancePostingRule.PAYMENT,
                FinancePostingRule.REMITTANCE,
                FinancePostingRule.CANCELLATION,
                FinancePostingRule.REPLACEMENT,
                FinancePostingRule.REVERSAL,
            }
            if not event_kinds.intersection({FinancePostingRule.RECOGNITION, FinancePostingRule.LIQUIDATION}):
                raise ValidationError(f"{variant.label} needs a reviewed recognition or liquidation rule.")
            missing = required_payment_events - event_kinds
            if missing:
                labels = dict(FinancePostingRule.EVENT_KIND_CHOICES)
                raise ValidationError(
                    f"{variant.label} still needs payment-cycle decisions for: "
                    + ", ".join(labels[kind] for kind in sorted(missing))
                    + ". Use an explicit no-entry rule where locally accepted."
                )
            posting_rules = list(variant.posting_rules.all())
            policy_error = payment_event_policy_error(posting_rules)
            if policy_error:
                raise ValidationError(f"{variant.label}: {policy_error}")
            for posting_rule in posting_rules:
                if posting_rule.authority_reference.startswith("EDIT BEFORE SUBMISSION"):
                    raise ValidationError(
                        f"{posting_rule.title} is still an editable starter. Replace its authority note with the reviewed local basis."
                    )
                posting_rule_snapshot(posting_rule)
        release.status, release.submitted_by, release.submitted_at = "submitted", actor, timezone.now()
        release.items.filter(status="draft").update(status="submitted")
        release.templates.filter(status="draft").update(status="submitted")
        release.signatories.filter(status="draft").update(status="submitted")
        release.parties.filter(status="draft").update(status="submitted")
        release.numbering_sequences.filter(status="draft").update(status="submitted")
        release.transaction_variants.filter(status="draft").update(status="submitted")
        fields = ("status", "submitted_by", "submitted_at", "updated_at")
    elif action == "approve":
        if not can_approve_finance_configuration(actor, release.department):
            raise PermissionDenied
        if release.status != "submitted":
            raise ValidationError("Only submitted releases can be approved.")
        if actor.pk in {release.created_by_id, release.submitted_by_id}:
            from .exemptions import workflow_exemption_for, workflow_exemption_snapshot
            from .models import FinanceWorkflowExemption

            exemption = workflow_exemption_for(
                actor=actor,
                control_code=FinanceWorkflowExemption.RELEASE_SELF_APPROVAL,
                department_id=release.department_id,
            )
            if exemption is None:
                raise ValidationError(
                    "The approver must be different from the release preparer and submitter unless an active "
                    "administrator-authorized workflow exemption applies."
                )
            workflow_exemption = workflow_exemption_snapshot(exemption)
        if not reason.strip():
            raise ValidationError("Record the local Accounting approval basis before approval.")
        failed_templates = release.templates.exclude(preflighted_at__isnull=False, preflight_result__passed=True)
        if failed_templates.exists():
            raise ValidationError("Every workbook template in the release must pass preflight before approval.")
        try:
            for template in release.templates.all():
                verify_template_evidence(template)
        except FinanceTemplateError as exc:
            raise ValidationError(str(exc)) from exc
        release.status, release.approved_by, release.approved_at = "approved", actor, timezone.now()
        release.items.filter(status="submitted").update(status="approved")
        release.templates.filter(status="submitted").update(status="approved")
        release.signatories.filter(status="submitted").update(status="approved")
        release.parties.filter(status="submitted").update(status="approved")
        for party in release.parties.all():
            party.authorized_claimants.filter(status="draft").update(status="approved")
        release.numbering_sequences.filter(status="submitted").update(status="approved")
        release.transaction_variants.filter(status="submitted").update(status="approved")
        release.accounting_approval_note = reason.strip()
        fields = ("status", "approved_by", "approved_at", "accounting_approval_note", "updated_at")
    elif action == "activate":
        if not can_approve_finance_configuration(actor, release.department):
            raise PermissionDenied
        if release.status not in {"approved", "scheduled"}:
            raise ValidationError("Only approved or scheduled releases can be activated.")
        if release.effective_from > timezone.localdate():
            raise ValidationError("A future-dated release must be scheduled and cannot activate early.")
        if release.effective_to and release.effective_to < timezone.localdate():
            raise ValidationError("An expired configuration release cannot be activated.")
        try:
            for template in release.templates.all():
                verify_template_evidence(template)
        except FinanceTemplateError as exc:
            raise ValidationError(str(exc)) from exc
        readiness = evaluate_readiness(release, as_of=release.effective_from)
        if not readiness["ready"]:
            raise ValidationError("Activation is blocked: " + "; ".join(item["message"] for item in readiness["blocking"]))
        preceding = list(FinanceConfigurationRelease.objects.select_for_update().filter(
            department=release.department, status="active"
        ).exclude(pk=release.pk))
        for prior in preceding:
            prior.status = "superseded"
            prior.save(update_fields=("status", "updated_at"))
            prior.items.filter(status="active").update(status="superseded")
            prior.templates.filter(status="active").update(status="superseded")
            prior.signatories.filter(status="active").update(status="superseded")
            prior.parties.filter(status="active").update(status="superseded")
            for party in prior.parties.all():
                party.authorized_claimants.filter(status="active").update(status="superseded")
            prior.numbering_sequences.filter(status="active").update(status="superseded")
            prior.transaction_variants.filter(status="active").update(status="superseded")
            record_event(prior, actor, "superseded", f"Superseded by {release}.")
        release.status, release.activated_by, release.activated_at = "active", actor, timezone.now()
        release.items.filter(status__in=("approved", "scheduled")).update(status="active")
        release.templates.filter(status__in=("approved", "scheduled")).update(status="active")
        release.signatories.filter(status__in=("approved", "scheduled")).update(status="active")
        release.parties.filter(status__in=("approved", "scheduled")).update(status="active")
        for party in release.parties.all():
            party.authorized_claimants.filter(status__in=("approved", "scheduled")).update(status="active")
        release.numbering_sequences.filter(status__in=("approved", "scheduled")).update(status="active")
        release.transaction_variants.filter(status__in=("approved", "scheduled")).update(status="active")
        fields = ("status", "activated_by", "activated_at", "updated_at")
    elif action == "schedule":
        if not can_approve_finance_configuration(actor, release.department):
            raise PermissionDenied
        if release.status != "approved" or release.effective_from <= timezone.localdate():
            raise ValidationError("Scheduling requires an approved release with a future effective date.")
        release.status = "scheduled"
        release.items.filter(status="approved").update(status="scheduled")
        release.templates.filter(status="approved").update(status="scheduled")
        release.signatories.filter(status="approved").update(status="scheduled")
        release.parties.filter(status="approved").update(status="scheduled")
        release.numbering_sequences.filter(status="approved").update(status="scheduled")
        release.transaction_variants.filter(status="approved").update(status="scheduled")
        fields = ("status", "updated_at")
    elif action == "rollback":
        if not can_approve_finance_configuration(actor, release.department):
            raise PermissionDenied
        if release.status != "superseded":
            raise ValidationError("Only a previously superseded release can be restored.")
        if not reason.strip():
            raise ValidationError("Record the Accounting rollback basis.")
        readiness = evaluate_readiness(release, as_of=timezone.localdate())
        if not readiness["ready"]:
            raise ValidationError("Rollback is blocked: " + "; ".join(item["message"] for item in readiness["blocking"]))
        current = list(FinanceConfigurationRelease.objects.select_for_update().filter(department=release.department, status="active"))
        for prior in current:
            prior.status = "superseded"
            prior.save(update_fields=("status", "updated_at"))
            record_event(prior, actor, "superseded_by_rollback", f"Replaced by rollback to {release}.")
        release.status, release.activated_by, release.activated_at = "active", actor, timezone.now()
        release.items.filter(status="superseded").update(status="active")
        release.templates.filter(status="superseded").update(status="active")
        release.signatories.filter(status="superseded").update(status="active")
        release.parties.filter(status="superseded").update(status="active")
        for party in release.parties.all():
            party.authorized_claimants.filter(status="superseded").update(status="active")
        release.numbering_sequences.filter(status="superseded").update(status="active")
        release.transaction_variants.filter(status="superseded").update(status="active")
        fields = ("status", "activated_by", "activated_at", "updated_at")
    elif action == "retire":
        if not can_approve_finance_configuration(actor, release.department):
            raise PermissionDenied
        if release.status not in {"approved", "scheduled", "active", "superseded"}:
            raise ValidationError("This release cannot be retired from its current state.")
        release.status = "retired"
        fields = ("status", "updated_at")
    else:
        raise ValidationError("Unsupported finance release action.")
    release.save(update_fields=fields)
    record_event(release, actor, action, reason, evidence=workflow_exemption)
    return release


def evaluate_readiness(release, as_of=None):
    as_of = as_of or timezone.localdate()
    governed_statuses = ("approved", "scheduled", "active", "superseded")
    items = release.items.filter(status__in=governed_statuses, effective_from__lte=as_of).filter(
        models_q_open_ended("effective_to", as_of)
    )
    categories = set(items.values_list("category", flat=True))
    typed_variants = release.transaction_variants.filter(
        status__in=governed_statuses, effective_from__lte=as_of,
    ).filter(models_q_open_ended("effective_to", as_of))
    typed_variant_ready = typed_variants.exists() and not typed_variants.filter(document_rules__isnull=True).exists()
    typed_posting_ready = typed_variants.exists()
    required_payment_events = {
        FinancePostingRule.PAYMENT,
        FinancePostingRule.REMITTANCE,
        FinancePostingRule.CANCELLATION,
        FinancePostingRule.REPLACEMENT,
        FinancePostingRule.REVERSAL,
    }
    if typed_posting_ready:
        try:
            for variant in typed_variants.prefetch_related("posting_rules__lines"):
                rules = list(variant.posting_rules.all())
                event_kinds = {rule.event_kind for rule in rules}
                has_initial_recognition = bool(
                    event_kinds & {FinancePostingRule.RECOGNITION, FinancePostingRule.LIQUIDATION}
                )
                if not has_initial_recognition or not required_payment_events.issubset(event_kinds):
                    typed_posting_ready = False
                    break
                if payment_event_policy_error(rules):
                    typed_posting_ready = False
                    break
                for rule in rules:
                    posting_rule_snapshot(rule)
        except ValidationError:
            typed_posting_ready = False
    checks = [
        ("approved_voucher_template", release.templates.filter(status__in=governed_statuses, preflighted_at__isnull=False, effective_from__lte=as_of).filter(models_q_open_ended("effective_to", as_of)).exists(), "An approved, checksum-verified voucher template applies.", "No approved, preflighted voucher template applies."),
        (
            "transaction_type_checklist",
            typed_variant_ready or ("transaction_type" in categories and "document_requirement" in categories),
            "An approved transaction variant and supporting-document checklist apply.",
            "At least one typed transaction variant with document rules, or a legacy transaction/document checklist, is required.",
        ),
        (
            "transaction_posting_rules",
            typed_posting_ready or not typed_variants.exists(),
            "Each approved typed transaction variant governs initial recognition and the payment, remittance, cancellation, and replacement events.",
            "Every typed transaction variant needs locally reviewed initial-recognition and payment-cycle rules; an explicit no-entry decision is allowed where locally accepted.",
        ),
        ("active_signatory", release.signatories.filter(status__in=governed_statuses, valid_from__lte=as_of).filter(models_q_open_ended("valid_to", as_of)).exists(), "An approved signatory assignment covers the applicable date.", "No approved signatory is valid for the applicable date."),
        ("fund_and_payment_account", "fund" in categories and bool(categories & {"bank_account", "payment_method"}), "An approved fund and payment account or method apply.", "An approved fund and payment account or method are required."),
        ("approved_tax_rule", "tax_rule" in categories, "An approved tax/deduction rule is available.", "No approved tax/deduction rule is available."),
        ("numbering_sequence", release.numbering_sequences.filter(status__in=governed_statuses, fiscal_year=release.fiscal_year).exists(), f"An approved numbering sequence covers fiscal year {release.fiscal_year}.", f"No approved numbering sequence exists for fiscal year {release.fiscal_year}."),
    ]
    conflicts = FinanceConfigurationRelease.objects.filter(
        department=release.department, status__in=("scheduled", "active"), effective_from__lte=release.effective_to or date.max,
    ).exclude(pk=release.pk).filter(models_q_open_ended("effective_to", release.effective_from)).exists()
    checks.append(("activation_date_conflict", not conflicts, "No scheduled or active release overlaps this effective period.", "The release effective dates overlap another scheduled or active release."))
    result = [{"code": code, "passed": passed, "message": success if passed else failure, "help_anchor": f"finance-readiness-{code}"} for code, passed, success, failure in checks]
    blocking = [item for item in result if not item["passed"]]
    return {"ready": not blocking, "checks": result, "blocking": blocking, "as_of": as_of.isoformat(), "sandbox_available": True}


def models_q_open_ended(field, value):
    from django.db.models import Q
    return Q(**{f"{field}__isnull": True}) | Q(**{f"{field}__gte": value})


def _workbook_bytes(template):
    template.workbook.open("rb")
    try:
        return template.workbook.read()
    finally:
        template.workbook.close()


def _destination(workbook, name):
    defined = workbook.defined_names.get(name)
    if not defined:
        raise FinanceTemplateError(f"Missing required workbook-level named range: {name}.")
    destinations = list(defined.destinations)
    if len(destinations) != 1:
        raise FinanceTemplateError(f"{name} must point to exactly one worksheet range.")
    sheet_name, coordinate = destinations[0]
    if sheet_name not in workbook.sheetnames:
        raise FinanceTemplateError(f"{name} points to a worksheet that does not exist.")
    return workbook[sheet_name], coordinate


def inspect_finance_workbook(payload, document_type="disbursement-voucher"):
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            members = archive.infolist()
            names = {member.filename.lower() for member in members}
            if sum(member.file_size for member in members) > 50 * 1024 * 1024 or any(member.file_size > 25 * 1024 * 1024 for member in members):
                raise FinanceTemplateError("The expanded workbook is too large to inspect safely.")
    except (zipfile.BadZipFile, OSError) as exc:
        raise FinanceTemplateError("The uploaded file is not a valid macro-free XLSX workbook.") from exc
    if any("vbaproject" in name or name.endswith(".bin") for name in names):
        raise FinanceTemplateError("Macro-enabled workbook content is not allowed.")
    if any(name.startswith("xl/externallinks/") for name in names):
        raise FinanceTemplateError("External workbook links are not allowed.")
    try:
        workbook = load_workbook(io.BytesIO(payload), data_only=False, keep_links=False)
    except Exception as exc:
        raise FinanceTemplateError("The XLSX workbook could not be opened safely.") from exc
    suspicious = ("WEBSERVICE(", "HYPERLINK(", "RTD(", "DDE(", "CALL(")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    normalized = cell.value.upper().replace(" ", "")
                    if "[" in cell.value or any(token in normalized for token in suspicious):
                        raise FinanceTemplateError(f"Suspicious or externally linked formula found in {sheet.title}!{cell.coordinate}.")
    schema = FinanceTemplateVersion.schema_for(document_type)
    if not schema:
        raise FinanceTemplateError("That finance document type does not have an approved controlled-range schema.")
    mapping = {}
    table_name = schema["table"]
    for name in schema["required"]:
        sheet, coordinate = _destination(workbook, name)
        if name != table_name:
            area = sheet[coordinate]
            if isinstance(area, tuple):
                flattened = [cell for row in area for cell in (row if isinstance(row, tuple) else (row,))]
                if len(flattened) != 1:
                    raise FinanceTemplateError(f"{name} must point to exactly one cell.")
        mapping[name] = {"worksheet": sheet.title, "range": coordinate}
    row_capacity = 0
    if table_name:
        line_sheet, line_coordinate = _destination(workbook, table_name)
        area = line_sheet[line_coordinate]
        if not isinstance(area, tuple):
            area = ((area,),)
        elif area and not isinstance(area[0], tuple):
            area = (area,)
        row_capacity = len(area)
        if row_capacity < 1:
            raise FinanceTemplateError(f"{table_name} must reserve at least one row.")
    print_sheets = [sheet.title for sheet in workbook.worksheets if sheet.print_area]
    if not print_sheets:
        raise FinanceTemplateError("Set a print area before submitting the voucher workbook.")
    mapped_sheets = {value["worksheet"] for value in mapping.values()}
    if mapped_sheets - set(print_sheets):
        raise FinanceTemplateError("Every worksheet receiving a controlled GRAND field must define a print area.")
    return workbook, mapping, {
        "passed": True, "worksheets": len(workbook.sheetnames), "required_names": len(mapping),
        "line_item_row_capacity": row_capacity, "print_area_sheets": print_sheets,
        "message": "Macro-free workbook, controlled names, formulas, print areas, and row capacity passed preflight.",
    }


@transaction.atomic
def preflight_finance_template(template, actor):
    if not can_manage_finance_templates(actor, template.department):
        raise PermissionDenied
    if template.status != "draft":
        raise ValidationError("Only draft template versions can be preflighted.")
    payload = _workbook_bytes(template)
    _workbook, mapping, result = inspect_finance_workbook(payload, template.document_type)
    template.workbook_checksum = hashlib.sha256(payload).hexdigest()
    template.mapping = mapping
    template.mapping_checksum = hashlib.sha256(json.dumps(mapping, sort_keys=True, separators=(",", ":")).encode()).hexdigest()
    template.preflight_result = result
    template.preflighted_by = actor
    template.preflighted_at = timezone.now()
    template.full_clean()
    template.save(update_fields=("workbook_checksum", "mapping", "mapping_checksum", "preflight_result", "preflighted_by", "preflighted_at"))
    record_event(template, actor, "preflight_passed")
    return result


def verify_template_evidence(template):
    payload = _workbook_bytes(template)
    if hashlib.sha256(payload).hexdigest() != template.workbook_checksum:
        raise FinanceTemplateError(f"{template} no longer matches its preflighted workbook checksum.")
    _workbook, mapping, _result = inspect_finance_workbook(payload, template.document_type)
    mapping_checksum = hashlib.sha256(json.dumps(mapping, sort_keys=True, separators=(",", ":")).encode()).hexdigest()
    if mapping_checksum != template.mapping_checksum:
        raise FinanceTemplateError(f"{template} no longer matches its preflighted named-range mapping.")
    return True


def synthetic_preview(template, actor):
    if not can_view_template_preview(actor, template):
        raise PermissionDenied
    payload = _workbook_bytes(template)
    checksum = hashlib.sha256(payload).hexdigest()
    if not template.preflight_passed or checksum != template.workbook_checksum:
        raise FinanceTemplateError("Preview requires the exact workbook version that passed preflight.")
    workbook, _mapping, _result = inspect_finance_workbook(payload, template.document_type)
    special_values = {
        "GRAND_DV_NUMBER": "SYNTHETIC-DV-000001", "GRAND_DV_DATE": timezone.localdate(),
        "GRAND_OBR_NUMBER": "SYNTHETIC-OBR-000001", "GRAND_OBR_DATE": timezone.localdate(),
        "GRAND_ADVICE_NUMBER": "SYNTHETIC-ADV-000001", "GRAND_ADVICE_DATE": timezone.localdate(),
        "GRAND_REGISTER_DATE": timezone.localdate(), "GRAND_RELEASE_DATE": timezone.localdate(),
        "GRAND_PAYEE": "Synthetic Demonstration Payee", "GRAND_PARTICULARS": "Synthetic preview only — not an official voucher",
        "GRAND_GROSS_AMOUNT": 1000, "GRAND_OBLIGATED_AMOUNT": 1000, "GRAND_TOTAL_DEDUCTIONS": 100, "GRAND_NET_AMOUNT": 900,
        "GRAND_BANK_ACCOUNT": "SYNTHETIC BANK ACCOUNT", "GRAND_CHECK_NUMBER": "SYNTHETIC-CHECK-000001",
        "GRAND_CLAIMANT": "Synthetic Authorized Claimant", "GRAND_FUND": "Synthetic Fund",
        "GRAND_RESPONSIBILITY_CENTER": "Synthetic Office", "GRAND_ACCOUNT_CODE": "SYNTHETIC-ACCOUNT",
        "GRAND_PREPARED_BY": "Sample Preparer", "GRAND_CERTIFIED_BY": "Sample Certifier", "GRAND_APPROVED_BY": "Sample Approver",
        "GRAND_RELEASED_BY": "Sample Releasing Officer", "GRAND_ACKNOWLEDGED_BY": "Sample Claimant",
    }
    schema = FinanceTemplateVersion.schema_for(template.document_type)
    table_name = schema["table"]
    values = {name: special_values.get(name, "SYNTHETIC PREVIEW") for name in schema["required"] if name != table_name}
    for name, value in values.items():
        sheet, coordinate = _destination(workbook, name)
        cells = sheet[coordinate]
        if isinstance(cells, tuple):
            cell = cells[0][0] if isinstance(cells[0], tuple) else cells[0]
        else:
            cell = cells
        cell.value = value
    if table_name:
        sheet, coordinate = _destination(workbook, table_name)
        cells = sheet[coordinate]
        if cells and not isinstance(cells[0], tuple):
            cells = (cells,)
        sample = ("Synthetic line / check", "SYNTHETIC-CODE", 1000, 900)
        for index, cell in enumerate(cells[0]):
            cell.value = sample[index] if index < len(sample) else None
    workbook.properties.title = "GRAND synthetic finance template preview"
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def can_view_template_preview(actor, template):
    from .access import can_view_finance_setup
    return can_view_finance_setup(actor, template.department)
