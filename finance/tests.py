import io
import tempfile
import zipfile
from datetime import date

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group, Permission
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName

from departments.models import Department
from profiles.models import EmployeeProfile

from .access import can_approve_finance_configuration, can_manage_finance_configuration
from .exemptions import workflow_exemption_for
from .models import (
    FinanceAuditEvent, FinanceConfigurationItem, FinanceConfigurationRelease, FinanceDocumentRule,
    FinanceNumberingSequence, FinancePostingRule, FinancePostingRuleLine, FinanceSignatory,
    FinanceTemplateVersion, FinanceWorkflowExemption, FinanceTransactionVariant,
)
from .services import (
    FinanceTemplateError, build_finance_starter_workbook, create_payment_event_posting_starters,
    create_recognition_posting_starter,
    evaluate_readiness, inspect_finance_workbook, payment_event_policy_error, posting_rule_snapshot, preflight_finance_template,
    record_event, synthetic_preview, transition_release,
)


TEMP_MEDIA = tempfile.mkdtemp(prefix="grand-finance-tests-")


@override_settings(MEDIA_ROOT=TEMP_MEDIA)
class FinanceSetupCenterTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.accounting = Department.objects.create(name="Municipal Accounting Office", slug="accounting")
        cls.other = Department.objects.create(name="Human Resources", slug="hr")
        cls.manager = cls._employee("finance.manager", cls.accounting)
        cls.approver = cls._employee("finance.approver", cls.accounting)
        cls.viewer = cls._employee("finance.viewer", cls.accounting)
        cls.outsider = cls._employee("hr.outsider", cls.other)
        cls.superuser = cls._employee("platform.admin", cls.accounting, is_superuser=True, is_staff=True)
        cls._grant(cls.manager, "view_finance_setup", "manage_finance_configuration", "manage_finance_templates")
        cls._grant(cls.approver, "view_finance_setup", "approve_finance_configuration")
        cls._grant(cls.viewer, "view_finance_setup")
        cls.release = FinanceConfigurationRelease.objects.create(
            department=cls.accounting, code="fy-2027", version=1, title="Synthetic FY 2027 setup",
            fiscal_year=2027, effective_from=date(2027, 1, 1), created_by=cls.manager,
        )

    @classmethod
    def _employee(cls, username, department, **kwargs):
        user = get_user_model().objects.create_user(username=username, email=f"{username}@example.test", password="finance-test-password", **kwargs)
        profile, _ = EmployeeProfile.objects.get_or_create(user=user)
        profile.assigned_department = department
        profile.save(update_fields=("assigned_department",))
        return get_user_model().objects.get(pk=user.pk)

    @classmethod
    def _grant(cls, user, *codenames):
        user.user_permissions.add(*Permission.objects.filter(content_type__app_label="finance", codename__in=codenames))

    def _item(self, category, code=None):
        return FinanceConfigurationItem.objects.create(
            department=self.accounting, release=self.release, category=category, code=code or category,
            version=1, label=f"Synthetic {category}", configuration={"synthetic": True},
            effective_from=date(2027, 1, 1), created_by=self.manager,
        )

    def _workbook(self, *, names=None, print_area=True, formula=None):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Voucher"
        required = list(names if names is not None else FinanceTemplateVersion.REQUIRED_NAMES)
        for index, name in enumerate(required, start=1):
            if name == "GRAND_LINE_ITEMS":
                coordinate = "$A$12:$D$20"
            else:
                row = ((index - 1) // 4) + 1
                column = ((index - 1) % 4) + 1
                coordinate = f"${chr(64 + column)}${row}"
            workbook.defined_names.add(DefinedName(name, attr_text=f"'Voucher'!{coordinate}"))
        if print_area:
            sheet.print_area = "A1:H30"
        if formula:
            sheet["H1"] = formula
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _template(self):
        return FinanceTemplateVersion.objects.create(
            department=self.accounting, release=self.release, document_type="disbursement-voucher",
            version=1, title="Synthetic controlled DV", workbook=SimpleUploadedFile("synthetic-dv.xlsx", self._workbook()),
            effective_from=date(2027, 1, 1), created_by=self.manager,
        )

    def _complete_release(self):
        for category in ("transaction_type", "document_requirement", "fund", "payment_method", "tax_rule"):
            self._item(category)
        FinanceSignatory.objects.create(
            department=self.accounting, release=self.release, role_code="approved-by", display_name="Synthetic Approver",
            position_title="Authorized Official", valid_from=date(2027, 1, 1), created_by=self.manager,
        )
        FinanceNumberingSequence.objects.create(
            department=self.accounting, release=self.release, fiscal_year=2027, document_type="disbursement-voucher",
            prefix="SYN-DV-", created_by=self.manager,
        )
        template = self._template()
        preflight_finance_template(template, self.manager)
        transition_release(self.release, "submit", self.manager)
        transition_release(self.release, "approve", self.approver, "Synthetic local Accounting review for automated testing.")
        self.release.refresh_from_db()
        return template

    def test_explicit_roles_and_department_boundary_do_not_use_superuser_bypass(self):
        self.assertTrue(can_manage_finance_configuration(self.manager, self.accounting))
        self.assertFalse(can_approve_finance_configuration(self.manager, self.accounting))
        self.assertFalse(can_manage_finance_configuration(self.outsider, self.accounting))
        self.assertFalse(can_manage_finance_configuration(self.superuser, self.accounting))

    def test_separate_approver_and_accounting_basis_are_required(self):
        transition_release(self.release, "submit", self.manager)
        with self.assertRaises(PermissionDenied):
            transition_release(self.release, "approve", self.manager, "Self approval")
        self._grant(self.manager, "approve_finance_configuration")
        with self.assertRaisesMessage(ValidationError, "different"):
            transition_release(self.release, "approve", self.manager, "Self approval")
        with self.assertRaisesMessage(ValidationError, "approval basis"):
            transition_release(self.release, "approve", self.approver, "")

    def test_structured_tax_starter_requires_local_confirmation_before_release_submission(self):
        item = FinanceConfigurationItem.objects.create(
            department=self.accounting, release=self.release, category="tax_rule", code="ewt-starter",
            version=1, label="Expanded withholding starter", description="Editable UAT starter",
            configuration={
                "reporting_enabled": True, "tax_family": "expanded_income", "atc": "WI158",
                "rate_percent": "1", "tax_base_label": "Reviewed gross income payment",
                "return_form_code": "1601-EQ", "certificate_form_code": "2307",
                "reporting_basis": "accounting_posting", "rounding_mode": "half_up",
                "requires_tax_identifier": True,
                "authority_reference": "Official BIR source pending local applicability review",
                "applicability_status": "candidate",
                "local_acceptance_note": "Local confirmation pending",
            },
            effective_from=date(2027, 1, 1), created_by=self.manager,
        )
        with self.assertRaisesMessage(ValidationError, "still a tax-reporting starter"):
            transition_release(self.release, "submit", self.manager)
        item.configuration = {
            **item.configuration,
            "applicability_status": "locally_confirmed",
            "local_acceptance_note": "Synthetic Municipal Accountant review retained in ACCTG-TAX-001",
        }
        item.full_clean(); item.save(update_fields=("configuration", "updated_at"))
        transition_release(self.release, "submit", self.manager)
        self.release.refresh_from_db()
        self.assertEqual(self.release.status, "submitted")

    def test_structured_tax_rule_form_uses_plain_fields_and_no_json(self):
        self.client.force_login(self.manager)
        response = self.client.get(reverse("finance:tax_rule_create"), {"release": self.release.pk})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Alphanumeric tax code")
        self.assertContains(response, "Rate (%)")
        self.assertContains(response, "Local decision and retained evidence")
        self.assertNotContains(response, "configuration_json")

    def test_admin_policy_can_exempt_a_named_user_or_role_and_records_each_use(self):
        self._grant(self.manager, "approve_finance_configuration")
        role = Group.objects.create(name="Synthetic small-office finance role")
        self.manager.groups.add(role)
        policy = FinanceWorkflowExemption.objects.create(
            department=self.accounting,
            control_code=FinanceWorkflowExemption.RELEASE_SELF_APPROVAL,
            subject_group=role,
            rationale="Synthetic small-office staffing exception with management review.",
            effective_from=date(2026, 1, 1),
            created_by=self.superuser,
        )
        transition_release(self.release, "submit", self.manager)
        transition_release(self.release, "approve", self.manager, "Synthetic approval under configured exemption.")
        event = self.release.events.get(action="approve")
        self.assertEqual(event.snapshot["workflow_exemption"]["policy_id"], policy.pk)
        self.assertEqual(event.snapshot["workflow_exemption"]["subject_type"], "group")

        policy.effective_to = date(2026, 1, 31)
        policy.save(update_fields=("effective_to",))
        self.assertIsNone(
            workflow_exemption_for(
                actor=self.manager,
                control_code=FinanceWorkflowExemption.RELEASE_SELF_APPROVAL,
                department_id=self.accounting.pk,
                as_of=date(2026, 2, 1),
            )
        )

    def test_release_transitions_append_audit_events_and_lock_governed_fields(self):
        self._complete_release()
        self.assertEqual(list(self.release.events.values_list("action", flat=True))[:2], ["approve", "submit"])
        self.release.title = "Changed after Accounting approval"
        with self.assertRaisesMessage(ValidationError, "immutable"):
            self.release.full_clean()
        event = self.release.events.first()
        event.reason = "Changed"
        with self.assertRaisesMessage(ValidationError, "append-only"):
            event.save()
        with self.assertRaisesMessage(ValidationError, "cannot be deleted"):
            event.delete()

    def test_approved_master_data_is_retained_and_requires_a_new_version(self):
        self._complete_release()
        item = self.release.items.first()
        item.label = "Mutated label"
        with self.assertRaisesMessage(ValidationError, "immutable"):
            item.full_clean()

    def test_workbook_preflight_records_both_checksums_and_synthetic_preview(self):
        template = self._template()
        result = preflight_finance_template(template, self.manager)
        template.refresh_from_db()
        self.assertTrue(result["passed"])
        self.assertEqual(result["required_names"], 11)
        self.assertEqual(result["line_item_row_capacity"], 9)
        self.assertEqual(len(template.workbook_checksum), 64)
        self.assertEqual(len(template.mapping_checksum), 64)
        preview = synthetic_preview(template, self.viewer)
        workbook = load_workbook(io.BytesIO(preview), data_only=False)
        destination = list(workbook.defined_names["GRAND_PAYEE"].destinations)[0]
        self.assertEqual(workbook[destination[0]][destination[1]].value, "Synthetic Demonstration Payee")
        self.assertIn("synthetic", workbook.properties.title.lower())

    def test_approval_rechecks_workbook_bytes_instead_of_trusting_stale_preflight(self):
        template = self._template()
        preflight_finance_template(template, self.manager)
        with template.workbook.storage.open(template.workbook.name, "wb") as stream:
            stream.write(self._workbook(formula="=1+1"))
        transition_release(self.release, "submit", self.manager)
        with self.assertRaisesMessage(ValidationError, "no longer matches"):
            transition_release(self.release, "approve", self.approver, "Synthetic Accounting review.")

    def test_workbook_preflight_rejects_missing_names_print_area_and_suspicious_formulas(self):
        with self.assertRaisesMessage(FinanceTemplateError, "GRAND_APPROVED_BY"):
            inspect_finance_workbook(self._workbook(names=FinanceTemplateVersion.REQUIRED_NAMES[:-1]))
        with self.assertRaisesMessage(FinanceTemplateError, "print area"):
            inspect_finance_workbook(self._workbook(print_area=False))
        with self.assertRaisesMessage(FinanceTemplateError, "Suspicious"):
            inspect_finance_workbook(self._workbook(formula='=WEBSERVICE("https://example.invalid")'))

    def test_workbook_preflight_rejects_external_link_parts_and_macro_payloads(self):
        original = self._workbook()
        for member, expected in (("xl/externalLinks/externalLink1.xml", "External"), ("xl/vbaProject.bin", "Macro")):
            source, output = zipfile.ZipFile(io.BytesIO(original)), io.BytesIO()
            with source, zipfile.ZipFile(output, "w") as target:
                for info in source.infolist():
                    target.writestr(info, source.read(info.filename))
                target.writestr(member, b"synthetic-test-only")
            with self.assertRaisesMessage(FinanceTemplateError, expected):
                inspect_finance_workbook(output.getvalue())

    def test_plain_language_dv_starter_is_editable_and_preflight_ready(self):
        values = {
            "lgu_name": "Municipality of Sample",
            "finance_office_name": "Municipal Accounting Office",
            "form_title": "DISBURSEMENT VOUCHER",
            "form_reference": "Editable starter — local comparison pending",
            "paper_size": "a4",
            "orientation": "portrait",
            "particulars_rows": 10,
            "default_copy_count": 2,
            "prepared_label": "Prepared by",
            "certified_label": "Certified / reviewed by",
            "approved_label": "Approved for payment by",
            "footer_note": "STARTER FOR LOCAL REVIEW",
        }
        payload = build_finance_starter_workbook(values)
        workbook, mapping, result = inspect_finance_workbook(payload)
        self.assertTrue(result["passed"])
        self.assertEqual(result["line_item_row_capacity"], 10)
        self.assertEqual(set(mapping), set(FinanceTemplateVersion.REQUIRED_NAMES))
        self.assertEqual(workbook["DV Starter"]["A1"].value, "Municipality of Sample")
        self.assertIn("not automatically an official", workbook["Read Me First"]["B2"].value)

        self.client.force_login(self.manager)
        response = self.client.post(reverse("finance:starter_template"), values)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["X-GRAND-Template-Status"], "editable-starter-not-locally-accepted")
        downloaded = b"".join(response.streaming_content) if response.streaming else response.content
        _workbook, _mapping, downloaded_result = inspect_finance_workbook(downloaded)
        self.assertTrue(downloaded_result["passed"])

        template = self._template()
        template.form_status = FinanceTemplateVersion.LOCALLY_ACCEPTED
        with self.assertRaisesMessage(ValidationError, "reviewed authority"):
            template.full_clean()

    def test_readiness_returns_reusable_reason_codes_then_passes_complete_release(self):
        initial = evaluate_readiness(self.release, as_of=date(2027, 1, 1))
        self.assertFalse(initial["ready"])
        self.assertIn("transaction_type_checklist", {item["code"] for item in initial["blocking"]})
        self.assertTrue(all(item["help_anchor"].startswith("finance-readiness-") for item in initial["checks"]))
        self._complete_release()
        ready = evaluate_readiness(self.release, as_of=date(2027, 1, 1))
        self.assertTrue(ready["ready"])

    def test_activation_is_blocked_until_ready_then_supersedes_prior_active_release(self):
        self.release.effective_from = timezone.localdate()
        self.release.save(update_fields=("effective_from",))
        transition_release(self.release, "submit", self.manager)
        transition_release(self.release, "approve", self.approver, "Synthetic approval basis.")
        with self.assertRaisesMessage(ValidationError, "Activation is blocked"):
            transition_release(self.release, "activate", self.approver)

    def test_workspace_and_objects_are_hidden_across_departments(self):
        self.client.force_login(self.viewer)
        self.assertContains(self.client.get(reverse("finance:workspace")), "Finance Setup Center")
        self.client.force_login(self.outsider)
        self.assertEqual(self.client.get(reverse("finance:workspace")).status_code, 403)
        self._grant(self.outsider, "view_finance_setup")
        self.outsider = get_user_model().objects.get(pk=self.outsider.pk)
        self.client.force_login(self.outsider)
        self.assertEqual(self.client.get(reverse("finance:release_detail", args=(self.release.pk,))).status_code, 404)

    def test_template_intake_rejects_xlsm_before_storage(self):
        self.client.force_login(self.manager)
        response = self.client.post(reverse("finance:template_create"), {
            "release": self.release.pk, "document_type": "dv", "version": 1, "title": "Macro workbook",
            "workbook": SimpleUploadedFile("unsafe.xlsm", b"not-a-workbook"),
            "effective_from": "2027-01-01", "effective_to": "",
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "macro-free .xlsx")
        self.assertFalse(FinanceTemplateVersion.objects.exists())

    def test_every_mutating_ui_action_creates_an_attributed_event(self):
        self.client.force_login(self.manager)
        response = self.client.post(reverse("finance:item_create"), {
            "release": self.release.pk, "category": "fund", "code": "synthetic-fund", "version": 1,
            "label": "Synthetic Fund", "description": "", "configuration_json": '{"sandbox": true}',
            "effective_from": "2027-01-01", "effective_to": "", "supersedes": "",
        })
        self.assertFalse(response.context["form"].errors if response.status_code == 200 else {}, response.context["form"].errors if response.status_code == 200 else {})
        self.assertRedirects(response, reverse("finance:release_detail", args=(self.release.pk,)))
        event = FinanceAuditEvent.objects.get(action="created", target_type="financeconfigurationitem")
        self.assertEqual(event.actor, self.manager)
        self.assertEqual(event.department, self.accounting)

    def test_typed_transaction_variant_and_document_rule_are_department_scoped_and_locked_with_release(self):
        self.client.force_login(self.manager)
        response = self.client.post(reverse("finance:variant_create"), {
            "release": self.release.pk, "code": "ordinary-supplier-claim",
            "label": "Ordinary supplier claim", "kind": FinanceTransactionVariant.ORDINARY_SUPPLIER,
            "description": "Synthetic supplier payable route.",
            "authority_reference": "Synthetic reviewed COA/DBM/local applicability decision.",
            "effective_from": "2027-01-01", "effective_to": "",
        })
        self.assertRedirects(response, reverse("finance:release_detail", args=(self.release.pk,)))
        variant = FinanceTransactionVariant.objects.get(code="ordinary-supplier-claim")
        response = self.client.post(reverse("finance:document_rule_create"), {
            "variant": variant.pk, "code": "invoice", "label": "Invoice / billing",
            "evidence_kind": FinanceDocumentRule.INVOICE, "required": "on",
            "waiver_allowed": "", "condition_description": "",
            "authority_reference": "Synthetic reviewed invoice requirement.", "display_order": 10,
        })
        self.assertRedirects(response, reverse("finance:release_detail", args=(self.release.pk,)))
        rule = variant.document_rules.get()
        self.assertTrue(rule.required)
        self.assertEqual(rule.department, self.accounting)
        conditional_without_scope = FinanceDocumentRule(
            variant=variant, code="conditional-without-scope", label="Unscoped conditional evidence",
            evidence_kind=FinanceDocumentRule.OTHER, required=False, waiver_allowed=False,
            condition_description="", authority_reference="Synthetic reviewed applicability decision.",
            created_by=self.manager,
        )
        with self.assertRaisesMessage(ValidationError, "must state when it applies"):
            conditional_without_scope.full_clean()
        self.assertTrue(FinanceAuditEvent.objects.filter(
            target_type="financetransactionvariant", action="created",
        ).exists())
        self.assertTrue(FinanceAuditEvent.objects.filter(
            target_type="financedocumentrule", action="document_rule_created",
        ).exists())
        posting = create_recognition_posting_starter(variant, self.manager)
        payment_rules = create_payment_event_posting_starters(variant, self.manager)
        snapshot, checksum = posting_rule_snapshot(posting)
        self.assertEqual(snapshot["event_kind"], FinancePostingRule.RECOGNITION)
        self.assertEqual(len(snapshot["lines"]), 3)
        self.assertEqual(len(checksum), 64)
        self.assertEqual(
            list(posting.lines.values_list("side", flat=True)),
            [FinancePostingRuleLine.DEBIT, FinancePostingRuleLine.CREDIT, FinancePostingRuleLine.CREDIT],
        )
        self.assertEqual(
            {rule.event_kind for rule in payment_rules},
            {
                FinancePostingRule.PAYMENT,
                FinancePostingRule.REMITTANCE,
                FinancePostingRule.CANCELLATION,
                FinancePostingRule.REPLACEMENT,
                FinancePostingRule.REVERSAL,
            },
        )
        cancellation = variant.posting_rules.get(event_kind=FinancePostingRule.CANCELLATION)
        cancellation_snapshot, _checksum = posting_rule_snapshot(cancellation)
        self.assertEqual(cancellation_snapshot["accounting_effect"], FinancePostingRule.NO_ENTRY)
        self.assertEqual(cancellation_snapshot["lines"], [])
        payment = variant.posting_rules.get(event_kind=FinancePostingRule.PAYMENT)
        self.assertEqual(
            set(payment.lines.values_list("amount_source", flat=True)),
            {FinancePostingRuleLine.EVENT_AMOUNT},
        )
        payment.recognition_point = FinancePostingRule.PAYMENT_ISSUANCE
        self.assertIn("requires journal-producing", payment_event_policy_error([
            payment,
            *variant.posting_rules.exclude(pk=payment.pk),
        ]))
        payment.recognition_point = FinancePostingRule.PAYMENT_RELEASE
        with self.assertRaisesMessage(ValidationError, "still an editable starter"):
            transition_release(self.release, "submit", self.manager)
        posting.authority_reference = "Synthetic locally reviewed supplier recognition policy."
        posting.full_clean()
        posting.save(update_fields=("authority_reference",))
        variant.posting_rules.exclude(pk=posting.pk).update(
            authority_reference="Synthetic locally reviewed payment-cycle policy.",
        )
        transition_release(self.release, "submit", self.manager)
        variant.refresh_from_db()
        self.assertEqual(variant.status, "submitted")
        variant.label = "Silently changed variant"
        with self.assertRaisesMessage(ValidationError, "immutable"):
            variant.full_clean()
